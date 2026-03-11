#!/usr/bin/env python3
"""
Music Royalty Valuation Tool - Desktop Version with AI-Powered Analysis
Double-click to run, select your earnings CSV, get a complete valuation spreadsheet.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
import re
import statistics

# Try to import AI analysis module
try:
    from ai_analysis import (
        run_full_analysis,
        GENRE_DECAY_BENCHMARKS,
        AIAnalysisResult
    )
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False


def create_valuation_template(royalty_name, year_minus_3, year_minus_2, year_minus_1, ytd, base_year,
                              output_path, ai_analysis=None, yearly_data=None, raw_data=None,
                              base_year_method=None, non_recurring_info=None, listing_price=0):
    """Creates the complete valuation template with data populated."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Model"

    # Define styles
    edit_font = Font(italic=True, color="0066CC")
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12)
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    scenario_bear_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    scenario_base_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    scenario_bull_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    weighted_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ai_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")

    # ============================================================================
    # TITLE
    # ============================================================================
    ws['A1'] = "MUSIC ROYALTY DCF VALUATION MODEL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Master Template with Weighted Scenario Analysis"
    ws['A2'].font = Font(italic=True, size=11, color="666666")

    # ============================================================================
    # DATA INPUT SECTION
    # ============================================================================
    ws['A4'] = "DATA INPUT"
    ws['A4'].font = section_font

    ws['A5'] = "Royalty Name/ID:"
    ws['B5'] = royalty_name
    ws['B5'].fill = input_fill
    ws['C5'] = "<- Edit"
    ws['C5'].font = edit_font

    ws['A6'] = "Listing Price"
    ws['B6'] = listing_price if listing_price > 0 else 0
    ws['B6'].fill = input_fill
    ws['B6'].number_format = '$#,##0.00'
    ws['C6'] = "<- Enter listing/bid price for IRR, payback & risk"
    ws['C6'].font = edit_font

    ws['A7'] = "HISTORICAL ROYALTIES"
    ws['A7'].font = header_font

    # Historical data - POPULATED
    ws['A8'] = "Year -3 Royalties"
    ws['B8'] = year_minus_3
    ws['B8'].fill = input_fill
    ws['B8'].number_format = '#,##0.00'
    ws['C8'] = "<- Edit"
    ws['C8'].font = edit_font

    ws['A9'] = "Year -2 Royalties"
    ws['B9'] = year_minus_2
    ws['B9'].fill = input_fill
    ws['B9'].number_format = '#,##0.00'
    ws['C9'] = "<- Edit"
    ws['C9'].font = edit_font

    ws['A10'] = "Year -1 Royalties"
    ws['B10'] = year_minus_1
    ws['B10'].fill = input_fill
    ws['B10'].number_format = '#,##0.00'
    ws['C10'] = "<- Edit"
    ws['C10'].font = edit_font

    ws['A11'] = "Current YTD Royalties"
    ws['B11'] = ytd
    ws['B11'].fill = input_fill
    ws['B11'].number_format = '#,##0.00'
    ws['C11'] = "<- Edit"
    ws['C11'].font = edit_font

    ws['A12'] = "3-Year Average"
    ws['B12'] = "=AVERAGE(B8:B10)"
    ws['B12'].number_format = '#,##0.00'

    ws['A13'] = "Base Year Royalties"
    ws['B13'] = base_year
    ws['B13'].fill = input_fill
    ws['B13'].number_format = '#,##0.00'
    base_year_note = "<- Edit (normalized starting CF)"
    if base_year_method:
        base_year_note += f"  [{base_year_method}]"
    ws['C13'] = base_year_note
    ws['C13'].font = edit_font

    # Non-recurring income note
    if non_recurring_info and non_recurring_info.get('amount', 0) > 0:
        nr_years = ', '.join(str(f['year']) for f in non_recurring_info.get('flagged_years', []))
        ws['D13'] = f"${non_recurring_info['amount']:,.0f} non-recurring removed from year(s) {nr_years}"
        ws['D13'].font = Font(italic=True, color="CC0000", size=9)

    # ============================================================================
    # KEY ASSUMPTIONS - With AI suggestions if available
    # ============================================================================
    ws['A15'] = "KEY ASSUMPTIONS"
    ws['A15'].font = section_font

    if ai_analysis:
        ws['D15'] = "AI SUGGESTED"
        ws['D15'].font = Font(bold=True, size=11, color="7B68EE")
        ws['E15'] = "Confidence"
        ws['E15'].font = Font(bold=True, size=11, color="7B68EE")

    ws['A16'] = "Growth Rate (Years 1-3)"
    ws['B16'] = 0.05
    ws['B16'].fill = input_fill
    ws['B16'].number_format = '0.0%'
    ws['C16'] = "<- Edit"
    ws['C16'].font = edit_font

    if ai_analysis:
        ws['D16'] = ai_analysis.suggested_growth_rate
        ws['D16'].fill = ai_fill
        ws['D16'].number_format = '0.0%'
        ws['E16'] = f"{ai_analysis.confidence_score:.0%}"
        ws['E16'].fill = ai_fill

    ws['A17'] = "Growth Rate (Years 4-5)"
    ws['B17'] = 0.03
    ws['B17'].fill = input_fill
    ws['B17'].number_format = '0.0%'
    ws['C17'] = "<- Edit"
    ws['C17'].font = edit_font

    if ai_analysis:
        ws['D17'] = ai_analysis.suggested_growth_rate * 0.6
        ws['D17'].fill = ai_fill
        ws['D17'].number_format = '0.0%'

    ws['A18'] = "Discount Rate"
    ws['B18'] = 0.12
    ws['B18'].fill = input_fill
    ws['B18'].number_format = '0.0%'
    ws['C18'] = "<- Edit"
    ws['C18'].font = edit_font

    if ai_analysis:
        ws['D18'] = ai_analysis.suggested_discount_rate
        ws['D18'].fill = ai_fill
        ws['D18'].number_format = '0.0%'

    ws['A19'] = "Terminal Growth Rate"
    ws['B19'] = -0.05
    ws['B19'].fill = input_fill
    ws['B19'].number_format = '0.0%'
    ws['C19'] = "<- Edit (usually negative)"
    ws['C19'].font = edit_font

    if ai_analysis:
        ws['D19'] = ai_analysis.suggested_terminal_rate
        ws['D19'].fill = ai_fill
        ws['D19'].number_format = '0.0%'

    # ============================================================================
    # SCENARIO ANALYSIS
    # ============================================================================
    ws['E4'] = "SCENARIO ANALYSIS"
    ws['E4'].font = section_font

    ws['F5'] = "Bear"
    ws['F5'].font = header_font
    ws['F5'].fill = scenario_bear_fill
    ws['F5'].alignment = Alignment(horizontal='center')
    ws['G5'] = "Base"
    ws['G5'].font = header_font
    ws['G5'].fill = scenario_base_fill
    ws['G5'].alignment = Alignment(horizontal='center')
    ws['H5'] = "Bull"
    ws['H5'].font = header_font
    ws['H5'].fill = scenario_bull_fill
    ws['H5'].alignment = Alignment(horizontal='center')

    # Scenario parameters
    ws['E6'] = "Base Year CF"
    ws['F6'] = "=B13*0.9"
    ws['G6'] = "=B13"
    ws['H6'] = "=B13*1.1"
    for col in ['F', 'G', 'H']:
        ws[f'{col}6'].number_format = '#,##0.00'

    ws['E7'] = "Growth (Yr 1-3)"
    ws['F7'] = "=B16-0.02"
    ws['G7'] = "=B16"
    ws['H7'] = "=B16+0.03"
    for col in ['F', 'G', 'H']:
        ws[f'{col}7'].number_format = '0.0%'

    ws['E8'] = "Growth (Yr 4-5)"
    ws['F8'] = "=B17-0.01"
    ws['G8'] = "=B17"
    ws['H8'] = "=B17+0.02"
    for col in ['F', 'G', 'H']:
        ws[f'{col}8'].number_format = '0.0%'

    ws['E9'] = "Discount Rate"
    ws['F9'] = "=B18+0.02"
    ws['G9'] = "=B18"
    ws['H9'] = "=B18"
    for col in ['F', 'G', 'H']:
        ws[f'{col}9'].number_format = '0.0%'

    ws['E10'] = "Terminal Growth"
    ws['F10'] = "=B19-0.02"
    ws['G10'] = "=B19"
    ws['H10'] = "=B19+0.02"
    for col in ['F', 'G', 'H']:
        ws[f'{col}10'].number_format = '0.0%'

    ws['E12'] = "Year 5 CF"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}12'] = f"={c}6*(1+{c}7)^3*(1+{c}8)^2"
        ws[f'{col}12'].number_format = '#,##0.00'

    ws['E13'] = "Terminal Value"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}13'] = f"=IFERROR({c}12*(1+{c}10)/MAX({c}9-{c}10,0.03),{c}12*10)"
        ws[f'{col}13'].number_format = '#,##0.00'

    ws['E14'] = "PV of Terminal"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}14'] = f"={c}13/(1+{c}9)^5"
        ws[f'{col}14'].number_format = '#,##0.00'

    ws['E16'] = "Implied Value"
    ws['E16'].font = header_font
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}16'] = (
            f"={c}6/(1+{c}9)"
            f"+{c}6*(1+{c}7)/(1+{c}9)^2"
            f"+{c}6*(1+{c}7)^2/(1+{c}9)^3"
            f"+{c}6*(1+{c}7)^3*(1+{c}8)/(1+{c}9)^4"
            f"+{c}12/(1+{c}9)^5"
            f"+{c}14"
        )
        ws[f'{col}16'].number_format = '$#,##0.00'
        ws[f'{col}16'].font = Font(bold=True)

    ws['E17'] = "vs Base Case"
    ws['F17'] = "=F16/G16-1"
    ws['G17'] = "-"
    ws['H17'] = "=H16/G16-1"
    for col in ['F', 'H']:
        ws[f'{col}17'].number_format = '0.0%'

    # ============================================================================
    # WEIGHTED AVERAGE VALUATION
    # ============================================================================
    ws['E19'] = "WEIGHTED AVERAGE VALUATION"
    ws['E19'].font = section_font

    ws['E20'] = "Scenario Weights"
    ws['E20'].font = header_font
    ws['F20'] = "Bear Weight"
    ws['G20'] = "Base Weight"
    ws['H20'] = "Bull Weight"

    # Use AI-suggested weights if available
    bear_weight = 0.25
    base_weight = 0.50
    bull_weight = 0.25
    if ai_analysis and ai_analysis.scenario_probabilities:
        bear_weight = ai_analysis.scenario_probabilities.get('bear', 25) / 100
        base_weight = ai_analysis.scenario_probabilities.get('base', 50) / 100
        bull_weight = ai_analysis.scenario_probabilities.get('bull', 25) / 100

    ws['F21'] = bear_weight
    ws['F21'].fill = input_fill
    ws['F21'].number_format = '0%'
    ws['G21'] = base_weight
    ws['G21'].fill = input_fill
    ws['G21'].number_format = '0%'
    ws['H21'] = bull_weight
    ws['H21'].fill = input_fill
    ws['H21'].number_format = '0%'
    ws['I21'] = "<- Edit weights (must = 100%)"
    ws['I21'].font = edit_font

    ws['E22'] = "Weight Check"
    ws['F22'] = "=F21+G21+H21"
    ws['F22'].number_format = '0%'
    ws['G22'] = '=IF(F22=1,"OK","ERROR: Must = 100%")'

    ws['E24'] = "WEIGHTED VALUATION"
    ws['E24'].font = Font(bold=True, size=12)
    ws['F24'] = "=F16*F21+G16*G21+H16*H21"
    ws['F24'].number_format = '$#,##0.00'
    ws['F24'].font = Font(bold=True, size=14)
    ws['F24'].fill = weighted_fill

    ws['E25'] = "Valuation Range"
    ws['F25'] = "=F16"
    ws['F25'].number_format = '$#,##0'
    ws['G25'] = "to"
    ws['H25'] = "=H16"
    ws['H25'].number_format = '$#,##0'

    ws['E26'] = "EV / Base Year CF"
    ws['F26'] = "=F24/B13"
    ws['F26'].number_format = '0.0x'

    # (TV dominance warning placed in Valuation Summary section below)

    # ============================================================================
    # INVESTOR METRICS (at listing price) — fully formula-based, reactive to B6
    # ============================================================================
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    warning_font = Font(bold=True, color="CC0000")

    ws['E29'] = "INVESTOR METRICS (at Listing Price)"
    ws['E29'].font = Font(bold=True, size=13)
    ws['E29'].fill = highlight_fill

    ws['E30'] = "Listing Price"
    ws['E30'].font = header_font
    ws['F30'] = "=B6"
    ws['F30'].number_format = '$#,##0'
    ws['F30'].fill = input_fill
    ws['G30'] = "<- Edit B6"
    ws['G30'].font = edit_font

    # Cash-on-cash yield (Problem 6)
    ws['E31'] = "Year 1 Cash Yield"
    ws['F31'] = '=IF(B6>0,C24/B6,"-")'
    ws['F31'].number_format = '0.0%'
    ws['G31'] = '=IF(AND(B6>0,C24/B6>0.08),"Strong yield",IF(AND(B6>0,C24/B6<0.04),"Low yield",""))'
    ws['G31'].font = Font(italic=True, color="666666", size=9)

    ws['E32'] = "Year 3 Cash Yield"
    ws['F32'] = '=IF(B6>0,E24/B6,"-")'
    ws['F32'].number_format = '0.0%'

    ws['E33'] = "Year 5 Cash Yield"
    ws['F33'] = '=IF(B6>0,G24/B6,"-")'
    ws['F33'].number_format = '0.0%'

    # Payback period (Problem 4)
    ws['E35'] = "Payback Period (years)"
    ws['E35'].font = header_font
    ws['F35'] = (
        '=IF(B6<=0,"-",'
        'IF(C24>=B6,1,'
        'IF(C24+D24>=B6,1+(B6-C24)/D24,'
        'IF(C24+D24+E24>=B6,2+(B6-C24-D24)/E24,'
        'IF(C24+D24+E24+F24>=B6,3+(B6-C24-D24-E24)/F24,'
        'IF(C24+D24+E24+F24+G24>=B6,4+(B6-C24-D24-E24-F24)/G24,'
        '"5+"))))))'
    )
    ws['F35'].number_format = '0.0'

    # IRR — uses helper rows with individual cell references (universal Excel compat)
    # The IRR() array constant syntax {-B6,...} fails in many Excel versions.
    # Instead: lay out cash flows in a hidden helper row, then IRR(range).

    # Helper rows for IRR cash flows — placed in columns J-O (out of main view)
    # Row 38: Base case CFs, Row 39: Bear, Row 40: Bull
    helper_col_start = 'J'  # J through O = 6 cells (yr0 through yr4+exit)
    helper_cols = ['J', 'K', 'L', 'M', 'N']  # 5 cells: -price, cf1, cf2, cf3, cf4+exit

    # Labels
    ws['J37'] = "IRR Helper (Base)"
    ws['J37'].font = Font(italic=True, color="999999", size=8)

    # --- Base case helper (row 38) ---
    ws['J38'] = "=-B6"                                          # -purchase price
    ws['K38'] = "=B13*(1+B16)"                                  # Year 1
    ws['L38'] = "=B13*(1+B16)^2"                                # Year 2
    ws['M38'] = "=B13*(1+B16)^3*(1+B17)"                       # Year 3 (transition to phase 2)
    # Year 4 + exit value (year 4 CF + PV of all future CFs as terminal)
    ws['N38'] = ("=B13*(1+B16)^3*(1+B17)^2"
                 "+B13*(1+B16)^3*(1+B17)^2*(1+B19)/MAX(B18-B19,0.03)/(1+B18)^5")

    # --- Bear case helper (row 39) ---
    ws['J39'] = "=-B6"
    ws['K39'] = "=B13*0.9*(1+(B16-0.02))"
    ws['L39'] = "=B13*0.9*(1+(B16-0.02))^2"
    ws['M39'] = "=B13*0.9*(1+(B16-0.02))^3*(1+(B17-0.01))"
    ws['N39'] = ("=B13*0.9*(1+(B16-0.02))^3*(1+(B17-0.01))^2"
                 "+B13*0.9*(1+(B16-0.02))^3*(1+(B17-0.01))^2*(1+(B19-0.02))/MAX((B18+0.02)-(B19-0.02),0.03)/(1+(B18+0.02))^5")

    # --- Bull case helper (row 40) ---
    ws['J40'] = "=-B6"
    ws['K40'] = "=B13*1.1*(1+(B16+0.03))"
    ws['L40'] = "=B13*1.1*(1+(B16+0.03))^2"
    ws['M40'] = "=B13*1.1*(1+(B16+0.03))^3*(1+(B17+0.02))"
    ws['N40'] = ("=B13*1.1*(1+(B16+0.03))^3*(1+(B17+0.02))^2"
                 "+B13*1.1*(1+(B16+0.03))^3*(1+(B17+0.02))^2*(1+(B19+0.02))/MAX(B18-(B19+0.02),0.03)/(1+B18)^5")

    # Hide helper columns
    for c in ['J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[c].hidden = True

    # --- Visible IRR section ---
    ws['E37'] = "IRR at Listing Price"
    ws['E37'].font = Font(bold=True, size=12)
    ws['E37'].fill = highlight_fill

    ws['E38'] = "Base Case IRR"
    ws['E38'].font = header_font
    ws['F38'] = '=IF(B6<=0,"-",IFERROR(IRR(J38:N38),"-"))'
    ws['F38'].number_format = '0.0%'
    ws['F38'].font = Font(bold=True, size=12)

    ws['E39'] = "Bear Case IRR"
    ws['F39'] = '=IF(B6<=0,"-",IFERROR(IRR(J39:N39),"-"))'
    ws['F39'].number_format = '0.0%'
    ws['F39'].fill = scenario_bear_fill

    ws['E40'] = "Bull Case IRR"
    ws['F40'] = '=IF(B6<=0,"-",IFERROR(IRR(J40:N40),"-"))'
    ws['F40'].number_format = '0.0%'
    ws['F40'].fill = scenario_bull_fill

    ws['E41'] = "Ref: 10yr Treasury ~4.3%, S&P div yield ~1.3%"
    ws['E41'].font = Font(italic=True, color="666666", size=9)

    # Breakeven decay rate (Problem 7) — Python-computed, written as value
    ws['E43'] = "MARGIN OF SAFETY"
    ws['E43'].font = Font(bold=True, size=12)
    ws['E43'].fill = highlight_fill

    if listing_price > 0 and base_year > 0:
        irr_data = compute_irr_and_breakeven(
            base_year_cf=base_year,
            growth_1_3=0.05, growth_4_5=0.03,
            discount_rate=0.12, terminal_growth=-0.05,
            listing_price=listing_price
        )
        if 'breakeven_decay' in irr_data:
            ws['E44'] = "Breakeven growth rate"
            ws['E44'].font = header_font
            ws['F44'] = irr_data['breakeven_decay']
            ws['F44'].number_format = '0.0%'
            ws['G44'] = "Growth rate where NPV = listing price"
            ws['G44'].font = Font(italic=True, color="666666", size=9)

            if 'margin_of_safety' in irr_data:
                ws['E45'] = "Margin of Safety"
                ws['E45'].font = header_font
                ws['F45'] = abs(irr_data['margin_of_safety'])
                ws['F45'].number_format = '0.0'
                ws['F45'].font = Font(bold=True, size=12)
                ws['G45'] = "percentage points of cushion"
                ws['G45'].font = Font(italic=True, color="666666", size=9)
        else:
            ws['E44'] = "Could not compute breakeven (enter listing price & re-run)"
            ws['E44'].font = Font(italic=True, color="666666")
    else:
        ws['E44'] = "Re-run tool with listing price to compute breakeven"
        ws['E44'].font = Font(italic=True, color="666666")

    # ============================================================================
    # 5-YEAR DCF PROJECTION
    # ============================================================================
    ws['A21'] = "5-YEAR DCF PROJECTION"
    ws['A21'].font = section_font

    headers = ["Year", "Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Terminal"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}22'] = h
        ws[f'{col}22'].font = header_font

    ws['A23'] = "Fiscal Year"
    ws['B23'] = datetime.now().year
    for i in range(1, 6):
        ws[f'{get_column_letter(i+2)}23'] = f"={get_column_letter(i+1)}23+1"
    ws['H23'] = "Perpetuity"

    ws['A24'] = "Royalty Income"
    ws['B24'] = "=B13"
    ws['C24'] = "=B24*(1+$B$16)"
    ws['D24'] = "=C24*(1+$B$16)"
    ws['E24'] = "=D24*(1+$B$16)"
    ws['F24'] = "=E24*(1+$B$17)"
    ws['G24'] = "=F24*(1+$B$17)"
    ws['H24'] = "=G24*(1+$B$19)"
    for col in 'BCDEFGH':
        ws[f'{col}24'].number_format = '#,##0.00'

    ws['A25'] = "Growth Rate"
    ws['B25'] = "-"
    ws['C25'] = "=$B$16"
    ws['D25'] = "=$B$16"
    ws['E25'] = "=$B$16"
    ws['F25'] = "=$B$17"
    ws['G25'] = "=$B$17"
    ws['H25'] = "=$B$19"
    for col in 'CDEFGH':
        ws[f'{col}25'].number_format = '0.0%'

    ws['A27'] = "Discount Factor"
    ws['B27'] = 1
    for i in range(1, 6):
        col = get_column_letter(i + 2)
        ws[f'{col}27'] = f"=1/(1+$B$18)^{i}"
        ws[f'{col}27'].number_format = '0.0000'
    ws['H27'] = "=G27"
    ws['H27'].number_format = '0.0000'

    ws['A28'] = "PV of Cash Flow"
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}28'] = f"={col}24*{col}27"
        ws[f'{col}28'].number_format = '#,##0.00'

    # ============================================================================
    # VALUATION SUMMARY
    # ============================================================================
    ws['A30'] = "VALUATION SUMMARY"
    ws['A30'].font = section_font

    ws['A31'] = "Terminal Value (undiscounted)"
    ws['B31'] = "=IFERROR(H24/MAX($B$18-$B$19,0.03),H24*10)"
    ws['B31'].number_format = '#,##0.00'
    ws['C31'] = "Gordon Growth (min 3% spread guard)"
    ws['C31'].font = Font(italic=True, color="666666")

    ws['A32'] = "PV of Terminal Value"
    ws['B32'] = "=B31*G27"
    ws['B32'].number_format = '#,##0.00'

    ws['A34'] = "Sum of PV of Cash Flows"
    ws['B34'] = "=SUM(C28:G28)"
    ws['B34'].number_format = '#,##0.00'

    ws['A35'] = "PV of Terminal Value"
    ws['B35'] = "=B32"
    ws['B35'].number_format = '#,##0.00'

    ws['A36'] = "Enterprise Value"
    ws['B36'] = "=B34+B35"
    ws['B36'].number_format = '$#,##0.00'
    ws['B36'].font = Font(bold=True)

    ws['A38'] = "% from Cash Flows"
    ws['B38'] = "=B34/B36"
    ws['B38'].number_format = '0.0%'

    ws['A39'] = "% from Terminal Value"
    ws['B39'] = "=B35/B36"
    ws['B39'].number_format = '0.0%'
    ws['C39'] = '=IF(B39>0.65,"Warning: TV dominates valuation — sensitive to long-term assumptions","")'
    ws['C39'].font = Font(italic=True, color="CC0000")

    # ============================================================================
    # SENSITIVITY ANALYSIS 1
    # ============================================================================
    # Sensitivity tables start row — pushed down to avoid investor metrics overlap
    sens1_start = 48

    ws[f'A{sens1_start}'] = "SENSITIVITY: Discount Rate vs Growth Rate (Years 1-3)"
    ws[f'A{sens1_start}'].font = section_font

    ws[f'A{sens1_start+1}'] = "Enterprise Value"
    ws[f'C{sens1_start+1}'] = "Growth Rate (Years 1-3)"
    ws[f'C{sens1_start+1}'].font = header_font

    growth_rates = [0.00, 0.02, 0.04, 0.06, 0.08, 0.10, 0.12]
    for i, gr in enumerate(growth_rates):
        col = get_column_letter(i + 3)
        ws[f'{col}{sens1_start+2}'] = gr
        ws[f'{col}{sens1_start+2}'].number_format = '0%'
        ws[f'{col}{sens1_start+2}'].font = header_font
        ws[f'{col}{sens1_start+2}'].alignment = Alignment(horizontal='center')

    ws[f'A{sens1_start+3}'] = "Discount"
    discount_rates = [0.08, 0.10, 0.12, 0.14, 0.16, 0.18]
    for i, dr in enumerate(discount_rates):
        row = sens1_start + 3 + i
        ws[f'B{row}'] = dr
        ws[f'B{row}'].number_format = '0%'
        ws[f'B{row}'].font = header_font

        for j in range(len(growth_rates)):
            col = get_column_letter(j + 3)
            formula = (
                f"=($B$13*(1+{col}${sens1_start+2})^3*(1+$B$17)^2*(1+$B$19)/MAX($B{row}-$B$19,0.03))/(1+$B{row})^5"
                f"+$B$13/(1+$B{row})"
                f"+$B$13*(1+{col}${sens1_start+2})/(1+$B{row})^2"
                f"+$B$13*(1+{col}${sens1_start+2})^2/(1+$B{row})^3"
                f"+$B$13*(1+{col}${sens1_start+2})^3*(1+$B$17)/(1+$B{row})^4"
                f"+$B$13*(1+{col}${sens1_start+2})^3*(1+$B$17)^2/(1+$B{row})^5"
            )
            ws[f'{col}{row}'] = formula
            ws[f'{col}{row}'].number_format = '#,##0'

    ws[f'A{sens1_start+4}'] = "Rate"

    # ============================================================================
    # SENSITIVITY ANALYSIS 2
    # ============================================================================
    sens2_start = sens1_start + 11

    ws[f'A{sens2_start}'] = "SENSITIVITY: Discount Rate vs Terminal Growth Rate"
    ws[f'A{sens2_start}'].font = section_font

    ws[f'A{sens2_start+1}'] = "Enterprise Value"
    ws[f'C{sens2_start+1}'] = "Terminal Growth Rate"
    ws[f'C{sens2_start+1}'].font = header_font

    term_growth_rates = [-0.10, -0.07, -0.05, -0.03, 0.00, 0.02, 0.03]
    for i, tg in enumerate(term_growth_rates):
        col = get_column_letter(i + 3)
        ws[f'{col}{sens2_start+2}'] = tg
        ws[f'{col}{sens2_start+2}'].number_format = '0%'
        ws[f'{col}{sens2_start+2}'].font = header_font
        ws[f'{col}{sens2_start+2}'].alignment = Alignment(horizontal='center')

    ws[f'A{sens2_start+3}'] = "Discount"
    for i, dr in enumerate(discount_rates):
        row = sens2_start + 3 + i
        ws[f'B{row}'] = dr
        ws[f'B{row}'].number_format = '0%'
        ws[f'B{row}'].font = header_font

        for j in range(len(term_growth_rates)):
            col = get_column_letter(j + 3)
            formula = (
                f"=($B$13*(1+$B$16)^3*(1+$B$17)^2*(1+{col}${sens2_start+2})/MAX($B{row}-{col}${sens2_start+2},0.03))/(1+$B{row})^5"
                f"+$B$13/(1+$B{row})"
                f"+$B$13*(1+$B$16)/(1+$B{row})^2"
                f"+$B$13*(1+$B$16)^2/(1+$B{row})^3"
                f"+$B$13*(1+$B$16)^3*(1+$B$17)/(1+$B{row})^4"
                f"+$B$13*(1+$B$16)^3*(1+$B$17)^2/(1+$B{row})^5"
            )
            ws[f'{col}{row}'] = formula
            ws[f'{col}{row}'].number_format = '#,##0'

    ws[f'A{sens2_start+4}'] = "Rate"

    # ============================================================================
    # MODEL NOTES
    # ============================================================================
    notes_start = sens2_start + 11
    ws[f'A{notes_start}'] = "MODEL NOTES"
    ws[f'A{notes_start}'].font = section_font

    notes = [
        "* Green cells are INPUT cells - edit these with your royalty data",
        "* Enter a Listing Price (B6) — IRR, payback, and yield update automatically",
        "* Purple cells show AI-suggested values (for reference)",
        "* Base Year is adaptively selected: stable=recent year, volatile=weighted avg or median",
        "* Non-recurring income (sync spikes, catch-ups) is detected and removed from base year",
        "* Terminal Value uses Gordon Growth with min 3% spread guard to prevent blowup",
        "* Breakeven decay rate requires re-running the tool with a listing price",
        "* Sensitivity tables show impact of key assumption changes"
    ]
    for i, note in enumerate(notes):
        ws[f'A{notes_start+1+i}'] = note
        ws[f'A{notes_start+1+i}'].font = Font(size=10, color="666666")

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 30

    # =========================================================================
    # AI INSIGHTS SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis:
        ws_ai = wb.create_sheet("AI Analysis")

        # Title
        ws_ai['A1'] = "AI-POWERED ANALYSIS"
        ws_ai['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_ai['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_ai['A2'].font = Font(italic=True, size=10, color="666666")

        # AI Narrative Summary
        ws_ai['A4'] = "EXECUTIVE SUMMARY"
        ws_ai['A4'].font = section_font
        ws_ai.merge_cells('A5:H7')
        ws_ai['A5'] = ai_analysis.ai_narrative
        ws_ai['A5'].alignment = Alignment(wrap_text=True, vertical='top')

        # Suggested Parameters Section
        ws_ai['A9'] = "AI-SUGGESTED PARAMETERS"
        ws_ai['A9'].font = section_font

        ws_ai['A10'] = "Parameter"
        ws_ai['B10'] = "AI Suggestion"
        ws_ai['C10'] = "Confidence"
        ws_ai['D10'] = "Rationale"
        for col in ['A', 'B', 'C', 'D']:
            ws_ai[f'{col}10'].font = header_font

        ws_ai['A11'] = "Growth Rate (Yr 1-3)"
        ws_ai['B11'] = ai_analysis.suggested_growth_rate
        ws_ai['B11'].number_format = '0.0%'
        ws_ai['B11'].fill = ai_fill
        ws_ai['C11'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D11'] = ai_analysis.haircut_rationale if ai_analysis.haircut_rationale else "Based on historical CAGR with mean reversion"

        ws_ai['A12'] = "Growth Rate (Yr 4-5)"
        ws_ai['B12'] = ai_analysis.suggested_growth_rate * 0.6
        ws_ai['B12'].number_format = '0.0%'
        ws_ai['B12'].fill = ai_fill
        ws_ai['C12'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D12'] = "Mature phase assumes slower growth"

        ws_ai['A13'] = "Terminal Growth Rate"
        ws_ai['B13'] = ai_analysis.suggested_terminal_rate
        ws_ai['B13'].number_format = '0.0%'
        ws_ai['B13'].fill = ai_fill
        ws_ai['C13'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D13'] = "Conservative perpetual decline assumption"

        ws_ai['A14'] = "Discount Rate"
        ws_ai['B14'] = ai_analysis.suggested_discount_rate
        ws_ai['B14'].number_format = '0.0%'
        ws_ai['B14'].fill = ai_fill
        ws_ai['C14'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D14'] = "Risk-adjusted based on volatility"

        # Growth Rate Methodology
        ws_ai['A16'] = "GROWTH RATE METHODOLOGY"
        ws_ai['A16'].font = section_font
        ws_ai['A17'] = "Recent CAGR (3yr)"
        ws_ai['B17'] = ai_analysis.raw_cagr
        ws_ai['B17'].number_format = '0.0%'
        ws_ai['A18'] = "Trend Direction"
        ws_ai['B18'] = ai_analysis.trend_direction.capitalize()
        ws_ai['A19'] = "Momentum"
        ws_ai['B19'] = ai_analysis.momentum.capitalize()
        ws_ai['A20'] = "Haircut Applied"
        ws_ai.merge_cells('B20:D20')
        ws_ai['B20'] = ai_analysis.haircut_rationale
        ws_ai['B20'].font = Font(italic=True, size=9)

        # Risk Factors
        ws_ai['A22'] = "RISK FACTORS"
        ws_ai['A22'].font = section_font
        for i, risk in enumerate(ai_analysis.risk_factors[:5]):
            ws_ai[f'A{23+i}'] = f"- {risk}"
            ws_ai[f'A{23+i}'].font = Font(color="CC0000")

        # Opportunities
        ws_ai['A29'] = "OPPORTUNITIES"
        ws_ai['A29'].font = section_font
        for i, opp in enumerate(ai_analysis.opportunities[:4]):
            ws_ai[f'A{30+i}'] = f"- {opp}"
            ws_ai[f'A{30+i}'].font = Font(color="006600")

        # Column widths
        ws_ai.column_dimensions['A'].width = 24
        ws_ai.column_dimensions['B'].width = 16
        ws_ai.column_dimensions['C'].width = 12
        ws_ai.column_dimensions['D'].width = 40

    # =========================================================================
    # MONTE CARLO SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis and ai_analysis.monte_carlo_results:
        ws_mc = wb.create_sheet("Monte Carlo")
        mc = ai_analysis.monte_carlo_results

        # Title
        ws_mc['A1'] = "MONTE CARLO SIMULATION RESULTS"
        ws_mc['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_mc['A2'] = f"Based on {mc['n_simulations']:,} simulations"
        ws_mc['A2'].font = Font(italic=True, size=10, color="666666")

        # Summary Statistics
        ws_mc['A4'] = "VALUATION DISTRIBUTION"
        ws_mc['A4'].font = section_font

        ws_mc['A5'] = "Statistic"
        ws_mc['B5'] = "Value"
        ws_mc['A5'].font = header_font
        ws_mc['B5'].font = header_font

        ws_mc['A6'] = "Mean Valuation"
        ws_mc['B6'] = mc['mean']
        ws_mc['B6'].number_format = '$#,##0'

        ws_mc['A7'] = "Median Valuation"
        ws_mc['B7'] = mc['median']
        ws_mc['B7'].number_format = '$#,##0'
        ws_mc['B7'].fill = weighted_fill
        ws_mc['B7'].font = Font(bold=True)

        ws_mc['A8'] = "Standard Deviation"
        ws_mc['B8'] = mc['std_dev']
        ws_mc['B8'].number_format = '$#,##0'

        # Percentiles
        ws_mc['A10'] = "PERCENTILE DISTRIBUTION"
        ws_mc['A10'].font = section_font

        percentiles = mc['percentiles']
        ws_mc['A11'] = "5th (Downside)"
        ws_mc['B11'] = percentiles['p5']
        ws_mc['B11'].number_format = '$#,##0'
        ws_mc['B11'].fill = scenario_bear_fill

        ws_mc['A12'] = "25th"
        ws_mc['B12'] = percentiles['p25']
        ws_mc['B12'].number_format = '$#,##0'

        ws_mc['A13'] = "50th (Median)"
        ws_mc['B13'] = percentiles['p50']
        ws_mc['B13'].number_format = '$#,##0'
        ws_mc['B13'].fill = scenario_base_fill
        ws_mc['B13'].font = Font(bold=True)

        ws_mc['A14'] = "75th"
        ws_mc['B14'] = percentiles['p75']
        ws_mc['B14'].number_format = '$#,##0'

        ws_mc['A15'] = "95th (Upside)"
        ws_mc['B15'] = percentiles['p95']
        ws_mc['B15'].number_format = '$#,##0'
        ws_mc['B15'].fill = scenario_bull_fill

        # Confidence Intervals
        ws_mc['A17'] = "CONFIDENCE INTERVALS"
        ws_mc['A17'].font = section_font

        ws_mc['A18'] = "90% Confidence Range:"
        ws_mc['B18'] = percentiles['p5']
        ws_mc['B18'].number_format = '$#,##0'
        ws_mc['C18'] = "to"
        ws_mc['D18'] = percentiles['p95']
        ws_mc['D18'].number_format = '$#,##0'

        ws_mc['A19'] = "80% Confidence Range:"
        ws_mc['B19'] = percentiles['p10']
        ws_mc['B19'].number_format = '$#,##0'
        ws_mc['C19'] = "to"
        ws_mc['D19'] = percentiles['p90']
        ws_mc['D19'].number_format = '$#,##0'

        # Write raw simulation valuations to a hidden sheet for formula-based risk metrics
        valuations = mc.get('valuations', [])
        n_sims = len(valuations)
        if n_sims > 0:
            ws_mc_data = wb.create_sheet("MC_Data")
            ws_mc_data.sheet_state = 'hidden'
            for idx, val in enumerate(valuations, start=1):
                ws_mc_data.cell(row=idx, column=1, value=val)
            data_range = f"MC_Data!$A$1:$A${n_sims}"

        # Risk metrics at listing price (Problem 8) — formula-driven from raw simulations
        ws_mc['A21'] = "RISK METRICS (at Listing Price)"
        ws_mc['A21'].font = section_font

        ws_mc['A22'] = "Listing Price"
        ws_mc['B22'] = "='Valuation Model'!B6"
        ws_mc['B22'].number_format = '$#,##0'

        if n_sims > 0:
            ws_mc['A23'] = "Probability of Loss"
            ws_mc['B23'] = f'=IF(\'Valuation Model\'!B6<=0,"-",COUNTIF({data_range},"<"&\'Valuation Model\'!B6)/{n_sims})'
            ws_mc['B23'].number_format = '0.0%'
            ws_mc['B23'].font = Font(bold=True)

            ws_mc['A24'] = "Value at Risk (95%)"
            ws_mc['B24'] = f'=IF(\'Valuation Model\'!B6<=0,"-",\'Valuation Model\'!B6-PERCENTILE({data_range},0.05))'
            ws_mc['B24'].number_format = '$#,##0'
            ws_mc['C24'] = "Max loss in 95% of scenarios"
            ws_mc['C24'].font = Font(italic=True, color="666666", size=9)

            ws_mc['A25'] = "Expected Shortfall (CVaR)"
            ws_mc['B25'] = f'=IF(\'Valuation Model\'!B6<=0,"-",\'Valuation Model\'!B6-AVERAGEIF({data_range},"<="&PERCENTILE({data_range},0.05)))'
            ws_mc['B25'].number_format = '$#,##0'
            ws_mc['C25'] = "Avg loss in worst 5% of scenarios"
            ws_mc['C25'].font = Font(italic=True, color="666666", size=9)
        else:
            ws_mc['A23'] = "No simulation data available"
            ws_mc['A23'].font = Font(italic=True, color="999999")

        # Column widths
        ws_mc.column_dimensions['A'].width = 26
        ws_mc.column_dimensions['B'].width = 16
        ws_mc.column_dimensions['C'].width = 32
        ws_mc.column_dimensions['D'].width = 14

    # =========================================================================
    # DECAY CURVE BENCHMARK SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis and ai_analysis.decay_curve_comparison:
        ws_decay = wb.create_sheet("Decay Benchmarks")
        benchmark = ai_analysis.decay_curve_comparison

        # Title
        ws_decay['A1'] = "DECAY CURVE BENCHMARK ANALYSIS"
        ws_decay['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_decay['A2'] = f"Genre: {benchmark.get('genre', 'Mixed').replace('_', ' ').title()}"
        ws_decay['A2'].font = Font(italic=True, size=11)

        # Overall Assessment
        ws_decay['A4'] = "OVERALL ASSESSMENT"
        ws_decay['A4'].font = section_font
        ws_decay['A5'] = benchmark.get('assessment', 'N/A')
        status = benchmark.get('overall_status', 'at_benchmark')
        if status == 'above_benchmark':
            ws_decay['A5'].font = Font(bold=True, color="006600")
        elif status == 'below_benchmark':
            ws_decay['A5'].font = Font(bold=True, color="CC0000")
        else:
            ws_decay['A5'].font = Font(bold=True)

        # Industry Benchmarks Reference
        ws_decay['A7'] = "INDUSTRY DECAY BENCHMARKS BY GENRE"
        ws_decay['A7'].font = section_font

        ws_decay['A8'] = "Genre"
        ws_decay['B8'] = "Year 1"
        ws_decay['C8'] = "Year 2"
        ws_decay['D8'] = "Year 3"
        ws_decay['E8'] = "Year 4"
        ws_decay['F8'] = "Year 5+"
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_decay[f'{col}8'].font = header_font

        row = 9
        for genre, rates in GENRE_DECAY_BENCHMARKS.items():
            ws_decay[f'A{row}'] = genre.replace('_', ' ').title()
            ws_decay[f'B{row}'] = rates['year_1']
            ws_decay[f'C{row}'] = rates['year_2']
            ws_decay[f'D{row}'] = rates['year_3']
            ws_decay[f'E{row}'] = rates['year_4']
            ws_decay[f'F{row}'] = rates['year_5_plus']
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_decay[f'{col}{row}'].number_format = '0%'

            # Highlight selected genre
            if genre == benchmark.get('genre', ''):
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_decay[f'{col}{row}'].fill = ai_fill
            row += 1

        # Column widths
        ws_decay.column_dimensions['A'].width = 16
        ws_decay.column_dimensions['B'].width = 10
        ws_decay.column_dimensions['C'].width = 10
        ws_decay.column_dimensions['D'].width = 10
        ws_decay.column_dimensions['E'].width = 10
        ws_decay.column_dimensions['F'].width = 10

    # =========================================================================
    # RAW DATA SHEET (if raw data provided)
    # =========================================================================
    if raw_data is not None:
        ws_raw = wb.create_sheet(title="Raw Data")
        for col_idx, col_name in enumerate(raw_data.columns, 1):
            cell = ws_raw.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for row_idx, row in enumerate(raw_data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_raw.cell(row=row_idx, column=col_idx, value=value)
        for col_idx, col_name in enumerate(raw_data.columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, min(len(raw_data) + 2, 100)):
                cell_value = ws_raw.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_raw.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    wb.save(output_path)
    return output_path


def detect_one_time_payments(df, year_col, amount_col, yearly):
    """
    Detect non-recurring income spikes using median + 2*MAD.
    Returns (adjusted_yearly, non_recurring_info).
    """
    values = yearly.values
    if len(values) < 3:
        return yearly, None

    med = float(statistics.median(values))
    deviations = [abs(v - med) for v in values]
    mad = float(statistics.median(deviations)) if len(deviations) > 0 else 0

    if mad == 0:
        return yearly, None

    threshold = med + 2 * mad
    flagged_years = []
    total_flagged = 0

    # Check for payment-type column (some CSVs have sync/mechanical/performance)
    type_col = None
    for col in df.columns:
        if col.lower() in ('type', 'payment_type', 'income_type', 'royalty_type', 'source_type'):
            type_col = col
            break

    adjusted_yearly = yearly.copy()
    for year in yearly.index:
        year_total = float(yearly[year])
        if year_total > threshold:
            excess = year_total - med
            reason = f"Year {int(year)}: ${excess:,.0f} above median (threshold: ${threshold:,.0f})"

            # If we have a type column, check for sync income
            if type_col is not None:
                year_mask = df[year_col] == year
                sync_mask = df[type_col].str.lower().str.contains('sync', na=False)
                sync_amount = float(df.loc[year_mask & sync_mask, amount_col].sum())
                if sync_amount > 0:
                    excess = sync_amount
                    reason = f"Year {int(year)}: ${sync_amount:,.0f} sync income identified"

            flagged_years.append({'year': int(year), 'amount': excess, 'reason': reason})
            total_flagged += excess
            adjusted_yearly[year] = year_total - excess

    if not flagged_years:
        return yearly, None

    return adjusted_yearly, {
        'amount': total_flagged,
        'original': float(yearly[yearly.index[-1]]) if len(yearly) > 0 else 0,
        'flagged_years': flagged_years,
        'reason': '; '.join(f['reason'] for f in flagged_years)
    }


def select_base_year(year_minus_1, year_minus_2, year_minus_3, ytd):
    """
    Adaptive base year selection based on coefficient of variation.
    Returns (base_year_value, method_description).
    """
    # Collect available full years
    full_years = [v for v in [year_minus_3, year_minus_2, year_minus_1] if v > 0]

    if len(full_years) == 0:
        return (ytd if ytd > 0 else 0), "YTD (no full years available)"

    if len(full_years) == 1:
        return full_years[0], "Most recent year (only 1 year available)"

    mean_val = statistics.mean(full_years)
    if mean_val == 0:
        return year_minus_1 if year_minus_1 > 0 else ytd, "Most recent year"

    std_val = statistics.stdev(full_years) if len(full_years) > 1 else 0
    cv = std_val / mean_val

    if cv < 0.15:
        # Stable: use most recent year
        return year_minus_1, f"Most recent year (CV={cv:.2f}, stable)"
    elif cv <= 0.40:
        # Moderate volatility: 3-year weighted average (50/30/20)
        if len(full_years) == 3:
            weighted = year_minus_1 * 0.50 + year_minus_2 * 0.30 + year_minus_3 * 0.20
        else:
            weighted = year_minus_1 * 0.60 + year_minus_2 * 0.40
        return weighted, f"Weighted avg (CV={cv:.2f}, moderate volatility)"
    else:
        # High volatility: 3-year median
        return statistics.median(full_years), f"Median (CV={cv:.2f}, high volatility)"


def compute_irr_and_breakeven(base_year_cf, growth_1_3, growth_4_5, discount_rate,
                               terminal_growth, listing_price):
    """
    Compute IRR at listing price and breakeven decay rate.
    Returns dict with irr_bear, irr_base, irr_bull, breakeven_decay, margin_of_safety.
    """
    result = {}

    if listing_price <= 0 or base_year_cf <= 0:
        return result

    # Project base-case cash flows
    def project_cfs(base_cf, g13, g45):
        cfs = [base_cf]
        for yr in range(1, 6):
            if yr <= 3:
                cfs.append(cfs[-1] * (1 + g13))
            else:
                cfs.append(cfs[-1] * (1 + g45))
        return cfs

    def compute_npv_at_growth(g, base_cf, dr, tg):
        """NPV with uniform growth rate g for years 1-5."""
        cfs = [base_cf]
        for yr in range(1, 6):
            cfs.append(cfs[-1] * (1 + g))
        pv_cfs = sum(cfs[i] / (1 + dr) ** i for i in range(1, 6))
        spread = max(dr - tg, 0.03)
        tv = cfs[5] * (1 + tg) / spread
        pv_tv = tv / (1 + dr) ** 5
        return pv_cfs + pv_tv

    # IRR computation
    try:
        import numpy_financial as npf

        scenarios = {
            'bear': (base_year_cf * 0.9, growth_1_3 - 0.02, growth_4_5 - 0.01,
                     discount_rate + 0.02, terminal_growth - 0.02),
            'base': (base_year_cf, growth_1_3, growth_4_5, discount_rate, terminal_growth),
            'bull': (base_year_cf * 1.1, growth_1_3 + 0.03, growth_4_5 + 0.02,
                     discount_rate, terminal_growth + 0.02),
        }

        for scenario_name, (bcf, g13, g45, dr, tg) in scenarios.items():
            cfs = project_cfs(bcf, g13, g45)
            # Terminal value as PV of remaining CFs after year 5
            spread = max(dr - tg, 0.03)
            tv = cfs[5] * (1 + tg) / spread
            pv_remaining = tv / (1 + dr) ** 5
            # IRR cash flow array: [-price, cf1, cf2, cf3, cf4, cf5 + pv_remaining_terminal]
            irr_cfs = [-listing_price] + cfs[1:5] + [cfs[5] + pv_remaining]
            try:
                irr_val = float(npf.irr(irr_cfs))
                if not (irr_val != irr_val):  # check NaN
                    result[f'irr_{scenario_name}'] = irr_val
            except Exception:
                pass
    except ImportError:
        pass

    # Breakeven decay rate (Problem 7)
    try:
        from scipy.optimize import brentq

        def npv_minus_price(g):
            return compute_npv_at_growth(g, base_year_cf, discount_rate, terminal_growth) - listing_price

        # Search for growth rate where NPV = listing price
        try:
            breakeven = brentq(npv_minus_price, -0.50, 0.50, xtol=1e-6)
            result['breakeven_decay'] = breakeven
            result['margin_of_safety'] = growth_1_3 - breakeven
        except ValueError:
            pass
    except ImportError:
        pass

    return result


def process_royalty_file(csv_path, genre="mixed", listing_price=0):
    """Process a royalty CSV file and create a valuation spreadsheet."""

    # Read the CSV
    df = pd.read_csv(csv_path)

    # Find the amount column
    amount_col = None
    for col in ['payable_amount', 'amount', 'earnings', 'royalty']:
        if col in df.columns.str.lower().tolist():
            amount_col = [c for c in df.columns if c.lower() == col][0]
            break

    if amount_col is None:
        # Try to find any column with 'amount' in the name
        amount_cols = [c for c in df.columns if 'amount' in c.lower()]
        if amount_cols:
            amount_col = amount_cols[0]
        else:
            raise ValueError("Could not find an amount/earnings column in the CSV")

    # Find the year column
    year_col = None
    for col in ['distribution_year', 'year', 'date']:
        if col in df.columns.str.lower().tolist():
            year_col = [c for c in df.columns if c.lower() == col][0]
            break

    if year_col is None:
        raise ValueError("Could not find a year column in the CSV")

    # Sum by year
    yearly = df.groupby(year_col)[amount_col].sum().sort_index()

    # Get years
    current_year = datetime.now().year
    years_list = sorted(yearly.index)

    # Extract values
    ytd = yearly.get(current_year, 0)
    year_minus_1 = yearly.get(current_year - 1, 0)
    year_minus_2 = yearly.get(current_year - 2, 0)
    year_minus_3 = yearly.get(current_year - 3, 0)

    # If no current year data, shift to most recent available
    shifted = False
    if ytd == 0 and years_list:
        shifted = True
        latest = max(years_list)
        ytd = yearly.get(latest, 0)
        year_minus_1 = yearly.get(latest - 1, 0)
        year_minus_2 = yearly.get(latest - 2, 0)
        year_minus_3 = yearly.get(latest - 3, 0)

    # Detect one-time payments (Problem 2)
    adjusted_yearly, non_recurring_info = detect_one_time_payments(df, year_col, amount_col, yearly)

    # Determine reference years (same shift logic as original extraction)
    if shifted:
        latest = max(years_list)
        ref_ym1_yr, ref_ym2_yr, ref_ym3_yr = latest - 1, latest - 2, latest - 3
    else:
        ref_ym1_yr = current_year - 1
        ref_ym2_yr = current_year - 2
        ref_ym3_yr = current_year - 3

    adj_ym1 = float(adjusted_yearly.get(ref_ym1_yr, 0))
    adj_ym2 = float(adjusted_yearly.get(ref_ym2_yr, 0))
    adj_ym3 = float(adjusted_yearly.get(ref_ym3_yr, 0))

    # Adaptive base year selection (Problem 1)
    adj_ytd = float(adjusted_yearly.get(max(years_list), 0)) if years_list else float(ytd)
    base_year, base_year_method = select_base_year(adj_ym1, adj_ym2, adj_ym3, adj_ytd)

    # Update non_recurring_info with the original base year value for display
    if non_recurring_info:
        non_recurring_info['original'] = float(year_minus_1) if year_minus_1 > 0 else float(ytd)

    # Convert yearly data to dict for AI analysis — use adjusted values for trend computation
    yearly_data_raw = {int(k): float(v) for k, v in yearly.items()}
    yearly_data_adjusted = {int(k): float(v) for k, v in adjusted_yearly.items()}

    # Run AI analysis on normalized earnings so growth/decay rates reflect true run-rate
    ai_analysis = None
    if AI_AVAILABLE and base_year > 0:
        try:
            ai_analysis = run_full_analysis(
                yearly_data=yearly_data_adjusted,
                base_year=base_year,
                genre=genre,
                n_simulations=1000,
                listing_price=listing_price
            )
        except Exception as e:
            print(f"AI analysis error: {e}")
            ai_analysis = None

    # Generate output filename
    base_name = os.path.basename(csv_path)
    if 'listing' in base_name.lower():
        match = re.search(r'listing[-_]?(\d+)', base_name, re.IGNORECASE)
        if match:
            royalty_name = f"Listing {match.group(1)}"
        else:
            royalty_name = os.path.splitext(base_name)[0]
    else:
        royalty_name = os.path.splitext(base_name)[0]

    # Save to "Output Sheets" folder within the tool's directory
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    output_dir = os.path.join(script_dir, "Output Sheets")
    os.makedirs(output_dir, exist_ok=True)

    output_filename = f"{royalty_name} Valuation.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # Create the valuation
    create_valuation_template(
        royalty_name=royalty_name,
        year_minus_3=year_minus_3,
        year_minus_2=year_minus_2,
        year_minus_1=year_minus_1,
        ytd=ytd,
        base_year=base_year,
        output_path=output_path,
        ai_analysis=ai_analysis,
        yearly_data=yearly_data_raw,
        raw_data=df,
        base_year_method=base_year_method,
        non_recurring_info=non_recurring_info,
        listing_price=listing_price
    )

    return output_path, royalty_name, yearly, ai_analysis


def main():
    # Hide the root window
    root = tk.Tk()
    root.withdraw()

    # Show file picker
    file_path = filedialog.askopenfilename(
        title="Select Royalty Earnings CSV",
        filetypes=[
            ("CSV files", "*.csv"),
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )

    if not file_path:
        messagebox.showinfo("Cancelled", "No file selected.")
        return

    try:
        output_path, royalty_name, yearly, ai_analysis = process_royalty_file(file_path)

        # Build summary message
        summary = f"Valuation created for: {royalty_name}\n\n"
        summary += "Yearly Royalties:\n"
        for year, amount in sorted(yearly.items()):
            summary += f"  {int(year)}: ${amount:,.2f}\n"
        summary += f"\nTotal: ${yearly.sum():,.2f}\n"

        if ai_analysis:
            summary += f"\n--- AI Analysis ---\n"
            summary += f"Suggested Growth: {ai_analysis.suggested_growth_rate:.1%}\n"
            summary += f"Suggested Discount: {ai_analysis.suggested_discount_rate:.1%}\n"
            summary += f"Confidence: {ai_analysis.confidence_score:.0%}\n"
            mc = ai_analysis.monte_carlo_results
            summary += f"\nMonte Carlo Median: ${mc['median']:,.0f}\n"
            summary += f"90% Range: ${mc['percentiles']['p5']:,.0f} - ${mc['percentiles']['p95']:,.0f}\n"

        summary += f"\nSaved to:\n{output_path}"

        messagebox.showinfo("Success!", summary)

        # Open the folder containing the file
        if sys.platform == 'darwin':  # macOS
            os.system(f'open -R "{output_path}"')
        elif sys.platform == 'win32':  # Windows
            os.system(f'explorer /select,"{output_path}"')

    except Exception as e:
        messagebox.showerror("Error", f"Failed to process file:\n\n{str(e)}")


if __name__ == "__main__":
    main()
