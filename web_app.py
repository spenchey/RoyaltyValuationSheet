#!/usr/bin/env python3
"""
Music Royalty Valuation Tool - Web Version with AI-Powered Analysis
Run this file and open the URL in any browser (including on your phone).
"""

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

from flask import Flask, request, send_file, render_template_string
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from datetime import datetime
import os
import io
import re

# Import AI analysis module
try:
    from ai_analysis import (
        run_full_analysis,
        GENRE_DECAY_BENCHMARKS,
        VALUATION_MULTIPLES,
        AIAnalysisResult
    )
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False

app = Flask(__name__)

# HTML Template - Mobile-friendly with AI options
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Royalty Valuation Tool</title>
    <style>
        * {
            box-sizing: border-box;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
        }
        body {
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }
        .container {
            max-width: 500px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            padding: 30px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
        }
        h1 {
            margin: 0 0 10px 0;
            color: #333;
            font-size: 24px;
            text-align: center;
        }
        .subtitle {
            color: #666;
            text-align: center;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .ai-badge {
            display: inline-block;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 11px;
            margin-left: 6px;
            vertical-align: middle;
        }
        .upload-area {
            border: 2px dashed #ddd;
            border-radius: 12px;
            padding: 40px 20px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
            margin-bottom: 20px;
        }
        .upload-area:hover {
            border-color: #667eea;
            background: #f8f9ff;
        }
        .upload-area.dragover {
            border-color: #667eea;
            background: #f0f3ff;
        }
        .upload-icon {
            font-size: 48px;
            margin-bottom: 10px;
        }
        .upload-text {
            color: #666;
            margin-bottom: 10px;
        }
        .upload-hint {
            color: #999;
            font-size: 12px;
        }
        input[type="file"] {
            display: none;
        }
        .file-name {
            background: #f0f3ff;
            padding: 12px 16px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: none;
            align-items: center;
            gap: 10px;
        }
        .file-name.show {
            display: flex;
        }
        .file-name span {
            flex: 1;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }
        .file-name button {
            background: none;
            border: none;
            color: #999;
            cursor: pointer;
            font-size: 18px;
        }
        .options-section {
            background: #f8f9ff;
            border-radius: 12px;
            padding: 16px;
            margin-bottom: 20px;
        }
        .options-title {
            font-weight: 600;
            font-size: 14px;
            margin-bottom: 12px;
            color: #333;
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .option-row {
            display: flex;
            align-items: center;
            margin-bottom: 12px;
        }
        .option-row:last-child {
            margin-bottom: 0;
        }
        .option-row label {
            flex: 1;
            font-size: 13px;
            color: #555;
        }
        .toggle {
            position: relative;
            width: 44px;
            height: 24px;
        }
        .toggle input {
            opacity: 0;
            width: 0;
            height: 0;
        }
        .toggle-slider {
            position: absolute;
            cursor: pointer;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background-color: #ccc;
            transition: .3s;
            border-radius: 24px;
        }
        .toggle-slider:before {
            position: absolute;
            content: "";
            height: 18px;
            width: 18px;
            left: 3px;
            bottom: 3px;
            background-color: white;
            transition: .3s;
            border-radius: 50%;
        }
        .toggle input:checked + .toggle-slider {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }
        .toggle input:checked + .toggle-slider:before {
            transform: translateX(20px);
        }
        select {
            padding: 8px 12px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 13px;
            background: white;
            min-width: 120px;
        }
        .api-key-input {
            width: 100%;
            padding: 10px 12px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 13px;
            margin-top: 8px;
        }
        .api-key-hint {
            font-size: 11px;
            color: #888;
            margin-top: 4px;
        }
        .submit-btn {
            width: 100%;
            padding: 16px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .submit-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
        }
        .submit-btn:disabled {
            background: #ccc;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }
        .error {
            background: #fee;
            color: #c00;
            padding: 12px 16px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: none;
        }
        .error.show {
            display: block;
        }
        .success {
            background: #efe;
            color: #060;
            padding: 12px 16px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: none;
        }
        .success.show {
            display: block;
        }
        .loading {
            display: none;
            text-align: center;
            padding: 20px;
        }
        .loading.show {
            display: block;
        }
        .spinner {
            width: 40px;
            height: 40px;
            border: 4px solid #f3f3f3;
            border-top: 4px solid #667eea;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin: 0 auto 10px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .instructions {
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #eee;
        }
        .instructions h3 {
            font-size: 14px;
            color: #333;
            margin: 0 0 10px 0;
        }
        .instructions ol {
            margin: 0;
            padding-left: 20px;
            color: #666;
            font-size: 13px;
        }
        .instructions li {
            margin-bottom: 6px;
        }
        .ai-features {
            margin-top: 16px;
            padding-top: 16px;
            border-top: 1px solid #e0e0e0;
        }
        .ai-features h4 {
            font-size: 13px;
            color: #667eea;
            margin: 0 0 8px 0;
        }
        .ai-features ul {
            margin: 0;
            padding-left: 18px;
            font-size: 12px;
            color: #666;
        }
        .ai-features li {
            margin-bottom: 4px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Royalty Valuation Tool <span class="ai-badge">AI</span></h1>
        <p class="subtitle">Upload your earnings CSV to generate a DCF valuation with AI-powered insights</p>

        <div class="error" id="error"></div>
        <div class="success" id="success"></div>

        <form id="uploadForm" action="/process" method="post" enctype="multipart/form-data">
            <div class="upload-area" id="uploadArea">
                <div class="upload-icon">📊</div>
                <div class="upload-text">Tap to select your CSV file</div>
                <div class="upload-hint">or drag and drop here</div>
            </div>
            <input type="file" name="file" id="fileInput" accept=".csv,.xlsx">

            <div class="file-name" id="fileName">
                <span id="fileNameText"></span>
                <button type="button" id="clearFile">&times;</button>
            </div>

            <div class="options-section">
                <div class="options-title">
                    <span>AI Analysis Options</span>
                </div>

                <div class="option-row">
                    <label for="enableAI">Enable AI-Powered Analysis</label>
                    <div class="toggle">
                        <input type="checkbox" id="enableAI" name="enable_ai" checked>
                        <span class="toggle-slider"></span>
                    </div>
                </div>

                <div class="option-row">
                    <label for="genre">Genre Classification</label>
                    <select id="genre" name="genre">
                        <option value="mixed">Mixed/Unknown</option>
                        <option value="pop">Pop</option>
                        <option value="rock">Rock</option>
                        <option value="hip_hop">Hip-Hop/Rap</option>
                        <option value="country">Country</option>
                        <option value="electronic">Electronic</option>
                        <option value="classical">Classical</option>
                    </select>
                </div>

                <div class="option-row">
                    <label for="simulations">Monte Carlo Simulations</label>
                    <select id="simulations" name="simulations">
                        <option value="500">500 (Fast)</option>
                        <option value="1000" selected>1,000 (Standard)</option>
                        <option value="5000">5,000 (Detailed)</option>
                    </select>
                </div>

                <div id="apiKeySection" style="margin-top: 12px;">
                    <label style="font-size: 13px; color: #555;">Claude OAuth Token (optional)</label>
                    {% if token_configured %}
                    <input type="password" class="api-key-input" id="oauthToken" name="oauth_token" placeholder="Using token from .env" style="background: #e8f5e9;">
                    <div class="api-key-hint" style="color: #2e7d32;">Token configured via .env file. Leave blank to use it, or enter a different one.</div>
                    {% else %}
                    <input type="password" class="api-key-input" id="oauthToken" name="oauth_token" placeholder="Enter your OAuth token...">
                    <div class="api-key-hint">For enhanced AI narrative. Works without it using statistical analysis.</div>
                    {% endif %}
                </div>
            </div>

            <div class="loading" id="loading">
                <div class="spinner"></div>
                <div id="loadingText">Generating valuation...</div>
            </div>

            <button type="submit" class="submit-btn" id="submitBtn" disabled>
                Generate Valuation
            </button>
        </form>

        <div class="instructions">
            <h3>How it works:</h3>
            <ol>
                <li>Upload your royalty earnings CSV</li>
                <li>We'll analyze the yearly totals</li>
                <li>AI generates projections &amp; risk analysis</li>
                <li>Download your complete DCF valuation spreadsheet</li>
            </ol>

            <div class="ai-features">
                <h4>AI-Powered Features:</h4>
                <ul>
                    <li>Monte Carlo simulations with probability distributions</li>
                    <li>Genre-based decay curve benchmarking</li>
                    <li>AI-suggested growth parameters</li>
                    <li>Risk factors and opportunity analysis</li>
                    <li>All original manual controls preserved</li>
                </ul>
            </div>
        </div>
    </div>

    <script>
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');
        const fileName = document.getElementById('fileName');
        const fileNameText = document.getElementById('fileNameText');
        const clearFile = document.getElementById('clearFile');
        const submitBtn = document.getElementById('submitBtn');
        const uploadForm = document.getElementById('uploadForm');
        const loading = document.getElementById('loading');
        const loadingText = document.getElementById('loadingText');
        const errorDiv = document.getElementById('error');
        const successDiv = document.getElementById('success');
        const enableAI = document.getElementById('enableAI');

        uploadArea.addEventListener('click', () => fileInput.click());

        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });

        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });

        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            if (e.dataTransfer.files.length) {
                fileInput.files = e.dataTransfer.files;
                updateFileName();
            }
        });

        fileInput.addEventListener('change', updateFileName);

        function updateFileName() {
            if (fileInput.files.length) {
                fileNameText.textContent = fileInput.files[0].name;
                fileName.classList.add('show');
                uploadArea.style.display = 'none';
                submitBtn.disabled = false;
                errorDiv.classList.remove('show');
            }
        }

        clearFile.addEventListener('click', () => {
            fileInput.value = '';
            fileName.classList.remove('show');
            uploadArea.style.display = 'block';
            submitBtn.disabled = true;
        });

        uploadForm.addEventListener('submit', async (e) => {
            e.preventDefault();

            loading.classList.add('show');
            submitBtn.disabled = true;
            errorDiv.classList.remove('show');
            successDiv.classList.remove('show');

            if (enableAI.checked) {
                loadingText.textContent = 'Running AI analysis & Monte Carlo simulations...';
            } else {
                loadingText.textContent = 'Generating valuation...';
            }

            const formData = new FormData(uploadForm);

            try {
                const response = await fetch('/process', {
                    method: 'POST',
                    body: formData
                });

                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = response.headers.get('X-Filename') || 'Valuation.xlsx';
                    document.body.appendChild(a);
                    a.click();
                    window.URL.revokeObjectURL(url);
                    a.remove();

                    successDiv.textContent = 'Valuation generated! Check your downloads.';
                    successDiv.classList.add('show');

                    // Reset form
                    fileInput.value = '';
                    fileName.classList.remove('show');
                    uploadArea.style.display = 'block';
                } else {
                    const text = await response.text();
                    throw new Error(text);
                }
            } catch (err) {
                errorDiv.textContent = err.message || 'Something went wrong. Please try again.';
                errorDiv.classList.add('show');
            } finally {
                loading.classList.remove('show');
                loadingText.textContent = 'Generating valuation...';
                submitBtn.disabled = !fileInput.files.length;
            }
        });
    </script>
</body>
</html>
"""


def create_valuation_template(royalty_name, year_minus_3, year_minus_2, year_minus_1, ytd, base_year,
                              ai_analysis=None, yearly_data=None, raw_data=None):
    """Creates the complete valuation template with data populated. Returns bytes."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Model"

    # Define styles
    edit_font = Font(italic=True, color="0066CC")
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=12)
    input_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    scenario_bear_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    scenario_base_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    scenario_bull_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    weighted_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ai_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")  # Light purple for AI suggestions

    # TITLE
    ws['A1'] = "MUSIC ROYALTY DCF VALUATION MODEL"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = "Master Template with Weighted Scenario Analysis"
    ws['A2'].font = Font(italic=True, size=11, color="666666")

    # DATA INPUT SECTION
    ws['A4'] = "DATA INPUT"
    ws['A4'].font = section_font

    ws['A5'] = "Royalty Name/ID:"
    ws['B5'] = royalty_name
    ws['B5'].fill = input_fill
    ws['C5'] = "<- Edit"
    ws['C5'].font = edit_font

    ws['A7'] = "HISTORICAL ROYALTIES"
    ws['A7'].font = header_font

    ws['A8'] = "Year -3 Royalties"
    ws['B8'] = year_minus_3
    ws['B8'].fill = input_fill
    ws['B8'].number_format = '#,##0.00'
    ws['C8'] = "<- Edit"
    ws['C8'].font = edit_font

    ws['A9'] = "Year -2 Royalties"
    ws['B9'] = year_minus_2
    ws['B9'].fill = input_fill
    ws['B9'].number_format = '#,##0.00'
    ws['C9'] = "<- Edit"
    ws['C9'].font = edit_font

    ws['A10'] = "Year -1 Royalties"
    ws['B10'] = year_minus_1
    ws['B10'].fill = input_fill
    ws['B10'].number_format = '#,##0.00'
    ws['C10'] = "<- Edit"
    ws['C10'].font = edit_font

    ws['A11'] = "Current YTD Royalties"
    ws['B11'] = ytd
    ws['B11'].fill = input_fill
    ws['B11'].number_format = '#,##0.00'
    ws['C11'] = "<- Edit"
    ws['C11'].font = edit_font

    ws['A12'] = "3-Year Average"
    ws['B12'] = "=AVERAGE(B8:B10)"
    ws['B12'].number_format = '#,##0.00'

    ws['A13'] = "Base Year Royalties"
    ws['B13'] = base_year
    ws['B13'].fill = input_fill
    ws['B13'].number_format = '#,##0.00'
    ws['C13'] = "<- Edit (normalized starting CF)"
    ws['C13'].font = edit_font

    # KEY ASSUMPTIONS - Now with AI suggestions shown alongside
    ws['A15'] = "KEY ASSUMPTIONS"
    ws['A15'].font = section_font

    # If AI analysis available, show suggestions
    if ai_analysis:
        ws['D15'] = "AI SUGGESTED"
        ws['D15'].font = Font(bold=True, size=11, color="7B68EE")
        ws['E15'] = "Confidence"
        ws['E15'].font = Font(bold=True, size=11, color="7B68EE")

    ws['A16'] = "Growth Rate (Years 1-3)"
    ws['B16'] = 0.05
    ws['B16'].fill = input_fill
    ws['B16'].number_format = '0.0%'
    ws['C16'] = "<- Edit"
    ws['C16'].font = edit_font

    if ai_analysis:
        ws['D16'] = ai_analysis.suggested_growth_rate
        ws['D16'].fill = ai_fill
        ws['D16'].number_format = '0.0%'
        ws['E16'] = f"{ai_analysis.confidence_score:.0%}"
        ws['E16'].fill = ai_fill

    ws['A17'] = "Growth Rate (Years 4-5)"
    ws['B17'] = 0.03
    ws['B17'].fill = input_fill
    ws['B17'].number_format = '0.0%'
    ws['C17'] = "<- Edit"
    ws['C17'].font = edit_font

    if ai_analysis:
        ws['D17'] = ai_analysis.suggested_growth_rate * 0.6
        ws['D17'].fill = ai_fill
        ws['D17'].number_format = '0.0%'

    ws['A18'] = "Discount Rate"
    ws['B18'] = 0.12
    ws['B18'].fill = input_fill
    ws['B18'].number_format = '0.0%'
    ws['C18'] = "<- Edit"
    ws['C18'].font = edit_font

    if ai_analysis:
        ws['D18'] = ai_analysis.suggested_discount_rate
        ws['D18'].fill = ai_fill
        ws['D18'].number_format = '0.0%'

    ws['A19'] = "Terminal Growth Rate"
    ws['B19'] = -0.05
    ws['B19'].fill = input_fill
    ws['B19'].number_format = '0.0%'
    ws['C19'] = "<- Edit (usually negative)"
    ws['C19'].font = edit_font

    if ai_analysis:
        ws['D19'] = ai_analysis.suggested_terminal_rate
        ws['D19'].fill = ai_fill
        ws['D19'].number_format = '0.0%'

    # SCENARIO ANALYSIS
    ws['E4'] = "SCENARIO ANALYSIS"
    ws['E4'].font = section_font

    ws['F5'] = "Bear"
    ws['F5'].font = header_font
    ws['F5'].fill = scenario_bear_fill
    ws['F5'].alignment = Alignment(horizontal='center')
    ws['G5'] = "Base"
    ws['G5'].font = header_font
    ws['G5'].fill = scenario_base_fill
    ws['G5'].alignment = Alignment(horizontal='center')
    ws['H5'] = "Bull"
    ws['H5'].font = header_font
    ws['H5'].fill = scenario_bull_fill
    ws['H5'].alignment = Alignment(horizontal='center')

    ws['E6'] = "Base Year CF"
    ws['F6'] = "=B13*0.9"
    ws['G6'] = "=B13"
    ws['H6'] = "=B13*1.1"
    for col in ['F', 'G', 'H']:
        ws[f'{col}6'].number_format = '#,##0.00'

    ws['E7'] = "Growth (Yr 1-3)"
    ws['F7'] = "=B16-0.02"
    ws['G7'] = "=B16"
    ws['H7'] = "=B16+0.03"
    for col in ['F', 'G', 'H']:
        ws[f'{col}7'].number_format = '0.0%'

    ws['E8'] = "Growth (Yr 4-5)"
    ws['F8'] = "=B17-0.01"
    ws['G8'] = "=B17"
    ws['H8'] = "=B17+0.02"
    for col in ['F', 'G', 'H']:
        ws[f'{col}8'].number_format = '0.0%'

    ws['E9'] = "Discount Rate"
    ws['F9'] = "=B18+0.02"
    ws['G9'] = "=B18"
    ws['H9'] = "=B18"
    for col in ['F', 'G', 'H']:
        ws[f'{col}9'].number_format = '0.0%'

    ws['E10'] = "Terminal Growth"
    ws['F10'] = "=B19-0.02"
    ws['G10'] = "=B19"
    ws['H10'] = "=B19+0.02"
    for col in ['F', 'G', 'H']:
        ws[f'{col}10'].number_format = '0.0%'

    ws['E12'] = "Year 5 CF"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}12'] = f"={c}6*(1+{c}7)^3*(1+{c}8)^2"
        ws[f'{col}12'].number_format = '#,##0.00'

    ws['E13'] = "Terminal Value"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}13'] = f"={c}12*(1+{c}10)/({c}9-{c}10)"
        ws[f'{col}13'].number_format = '#,##0.00'

    ws['E14'] = "PV of Terminal"
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}14'] = f"={c}13/(1+{c}9)^5"
        ws[f'{col}14'].number_format = '#,##0.00'

    ws['E16'] = "Implied Value"
    ws['E16'].font = header_font
    for col, c in [('F', 'F'), ('G', 'G'), ('H', 'H')]:
        ws[f'{col}16'] = (
            f"={c}6/(1+{c}9)"
            f"+{c}6*(1+{c}7)/(1+{c}9)^2"
            f"+{c}6*(1+{c}7)^2/(1+{c}9)^3"
            f"+{c}6*(1+{c}7)^3*(1+{c}8)/(1+{c}9)^4"
            f"+{c}12/(1+{c}9)^5"
            f"+{c}14"
        )
        ws[f'{col}16'].number_format = '$#,##0.00'
        ws[f'{col}16'].font = Font(bold=True)

    ws['E17'] = "vs Base Case"
    ws['F17'] = "=F16/G16-1"
    ws['G17'] = "-"
    ws['H17'] = "=H16/G16-1"
    for col in ['F', 'H']:
        ws[f'{col}17'].number_format = '0.0%'

    # WEIGHTED AVERAGE VALUATION
    ws['E19'] = "WEIGHTED AVERAGE VALUATION"
    ws['E19'].font = section_font

    ws['E20'] = "Scenario Weights"
    ws['E20'].font = header_font
    ws['F20'] = "Bear Weight"
    ws['G20'] = "Base Weight"
    ws['H20'] = "Bull Weight"

    # Use AI-suggested weights if available
    bear_weight = 0.25
    base_weight = 0.50
    bull_weight = 0.25
    if ai_analysis and ai_analysis.scenario_probabilities:
        bear_weight = ai_analysis.scenario_probabilities.get('bear', 25) / 100
        base_weight = ai_analysis.scenario_probabilities.get('base', 50) / 100
        bull_weight = ai_analysis.scenario_probabilities.get('bull', 25) / 100

    ws['F21'] = bear_weight
    ws['F21'].fill = input_fill
    ws['F21'].number_format = '0%'
    ws['G21'] = base_weight
    ws['G21'].fill = input_fill
    ws['G21'].number_format = '0%'
    ws['H21'] = bull_weight
    ws['H21'].fill = input_fill
    ws['H21'].number_format = '0%'
    ws['I21'] = "<- Edit weights (must = 100%)"
    ws['I21'].font = edit_font

    ws['E22'] = "Weight Check"
    ws['F22'] = "=F21+G21+H21"
    ws['F22'].number_format = '0%'
    ws['G22'] = '=IF(F22=1,"OK","ERROR: Must = 100%")'

    ws['E24'] = "WEIGHTED VALUATION"
    ws['E24'].font = Font(bold=True, size=12)
    ws['F24'] = "=F16*F21+G16*G21+H16*H21"
    ws['F24'].number_format = '$#,##0.00'
    ws['F24'].font = Font(bold=True, size=14)
    ws['F24'].fill = weighted_fill

    ws['E25'] = "Valuation Range"
    ws['F25'] = "=F16"
    ws['F25'].number_format = '$#,##0'
    ws['G25'] = "to"
    ws['H25'] = "=H16"
    ws['H25'].number_format = '$#,##0'

    ws['E26'] = "EV / Base Year CF"
    ws['F26'] = "=F24/B13"
    ws['F26'].number_format = '0.0x'

    ws['E27'] = "Payback Period (years)"
    ws['F27'] = "=F24/B13"
    ws['F27'].number_format = '0.0'

    # 5-YEAR DCF PROJECTION
    ws['A21'] = "5-YEAR DCF PROJECTION"
    ws['A21'].font = section_font

    headers = ["Year", "Base", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Terminal"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}22'] = h
        ws[f'{col}22'].font = header_font

    ws['A23'] = "Fiscal Year"
    ws['B23'] = datetime.now().year
    for i in range(1, 6):
        ws[f'{get_column_letter(i+2)}23'] = f"={get_column_letter(i+1)}23+1"
    ws['H23'] = "Perpetuity"

    ws['A24'] = "Royalty Income"
    ws['B24'] = "=B13"
    ws['C24'] = "=B24*(1+$B$16)"
    ws['D24'] = "=C24*(1+$B$16)"
    ws['E24'] = "=D24*(1+$B$16)"
    ws['F24'] = "=E24*(1+$B$17)"
    ws['G24'] = "=F24*(1+$B$17)"
    ws['H24'] = "=G24*(1+$B$19)"
    for col in 'BCDEFGH':
        ws[f'{col}24'].number_format = '#,##0.00'

    ws['A25'] = "Growth Rate"
    ws['B25'] = "-"
    ws['C25'] = "=$B$16"
    ws['D25'] = "=$B$16"
    ws['E25'] = "=$B$16"
    ws['F25'] = "=$B$17"
    ws['G25'] = "=$B$17"
    ws['H25'] = "=$B$19"
    for col in 'CDEFGH':
        ws[f'{col}25'].number_format = '0.0%'

    ws['A27'] = "Discount Factor"
    ws['B27'] = 1
    for i in range(1, 6):
        col = get_column_letter(i + 2)
        ws[f'{col}27'] = f"=1/(1+$B$18)^{i}"
        ws[f'{col}27'].number_format = '0.0000'
    ws['H27'] = "=G27"
    ws['H27'].number_format = '0.0000'

    ws['A28'] = "PV of Cash Flow"
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}28'] = f"={col}24*{col}27"
        ws[f'{col}28'].number_format = '#,##0.00'

    # VALUATION SUMMARY
    ws['A30'] = "VALUATION SUMMARY"
    ws['A30'].font = section_font

    ws['A31'] = "Terminal Value (undiscounted)"
    ws['B31'] = "=H24/($B$18-$B$19)"
    ws['B31'].number_format = '#,##0.00'
    ws['C31'] = "Gordon Growth formula"
    ws['C31'].font = Font(italic=True, color="666666")

    ws['A32'] = "PV of Terminal Value"
    ws['B32'] = "=B31*G27"
    ws['B32'].number_format = '#,##0.00'

    ws['A34'] = "Sum of PV of Cash Flows"
    ws['B34'] = "=SUM(C28:G28)"
    ws['B34'].number_format = '#,##0.00'

    ws['A35'] = "PV of Terminal Value"
    ws['B35'] = "=B32"
    ws['B35'].number_format = '#,##0.00'

    ws['A36'] = "Enterprise Value"
    ws['B36'] = "=B34+B35"
    ws['B36'].number_format = '$#,##0.00'
    ws['B36'].font = Font(bold=True)

    ws['A38'] = "% from Cash Flows"
    ws['B38'] = "=B34/B36"
    ws['B38'].number_format = '0.0%'

    ws['A39'] = "% from Terminal Value"
    ws['B39'] = "=B35/B36"
    ws['B39'].number_format = '0.0%'

    # SENSITIVITY ANALYSIS 1
    ws['A41'] = "SENSITIVITY: Discount Rate vs Growth Rate (Years 1-3)"
    ws['A41'].font = section_font

    ws['A42'] = "Enterprise Value"
    ws['C42'] = "Growth Rate (Years 1-3)"
    ws['C42'].font = header_font

    growth_rates = [0.00, 0.02, 0.04, 0.06, 0.08, 0.10, 0.12]
    for i, gr in enumerate(growth_rates):
        col = get_column_letter(i + 3)
        ws[f'{col}43'] = gr
        ws[f'{col}43'].number_format = '0%'
        ws[f'{col}43'].font = header_font
        ws[f'{col}43'].alignment = Alignment(horizontal='center')

    ws['A44'] = "Discount"
    discount_rates = [0.08, 0.10, 0.12, 0.14, 0.16, 0.18]
    for i, dr in enumerate(discount_rates):
        row = 44 + i
        ws[f'B{row}'] = dr
        ws[f'B{row}'].number_format = '0%'
        ws[f'B{row}'].font = header_font

        for j in range(len(growth_rates)):
            col = get_column_letter(j + 3)
            formula = (
                f"=($B$13*(1+{col}$43)^3*(1+$B$17)^2*(1+$B$19)/($B{row}-$B$19))/(1+$B{row})^5"
                f"+$B$13/(1+$B{row})"
                f"+$B$13*(1+{col}$43)/(1+$B{row})^2"
                f"+$B$13*(1+{col}$43)^2/(1+$B{row})^3"
                f"+$B$13*(1+{col}$43)^3*(1+$B$17)/(1+$B{row})^4"
                f"+$B$13*(1+{col}$43)^3*(1+$B$17)^2/(1+$B{row})^5"
            )
            ws[f'{col}{row}'] = formula
            ws[f'{col}{row}'].number_format = '#,##0'

    ws['A45'] = "Rate"

    # SENSITIVITY ANALYSIS 2
    ws['A52'] = "SENSITIVITY: Discount Rate vs Terminal Growth Rate"
    ws['A52'].font = section_font

    ws['A53'] = "Enterprise Value"
    ws['C53'] = "Terminal Growth Rate"
    ws['C53'].font = header_font

    term_growth_rates = [-0.10, -0.07, -0.05, -0.03, 0.00, 0.02, 0.03]
    for i, tg in enumerate(term_growth_rates):
        col = get_column_letter(i + 3)
        ws[f'{col}54'] = tg
        ws[f'{col}54'].number_format = '0%'
        ws[f'{col}54'].font = header_font
        ws[f'{col}54'].alignment = Alignment(horizontal='center')

    ws['A55'] = "Discount"
    for i, dr in enumerate(discount_rates):
        row = 55 + i
        ws[f'B{row}'] = dr
        ws[f'B{row}'].number_format = '0%'
        ws[f'B{row}'].font = header_font

        for j in range(len(term_growth_rates)):
            col = get_column_letter(j + 3)
            formula = (
                f"=($B$13*(1+$B$16)^3*(1+$B$17)^2*(1+{col}$54)/($B{row}-{col}$54))/(1+$B{row})^5"
                f"+$B$13/(1+$B{row})"
                f"+$B$13*(1+$B$16)/(1+$B{row})^2"
                f"+$B$13*(1+$B$16)^2/(1+$B{row})^3"
                f"+$B$13*(1+$B$16)^3*(1+$B$17)/(1+$B{row})^4"
                f"+$B$13*(1+$B$16)^3*(1+$B$17)^2/(1+$B{row})^5"
            )
            ws[f'{col}{row}'] = formula
            ws[f'{col}{row}'].number_format = '#,##0'

    ws['A56'] = "Rate"

    # MODEL NOTES
    ws['A62'] = "MODEL NOTES"
    ws['A62'].font = section_font

    notes = [
        "* Green cells are INPUT cells - edit these with your royalty data",
        "* Purple cells show AI-suggested values (for reference)",
        "* Royalties = pure cash flow (no costs modeled)",
        "* Terminal Value = Year 5 CF x (1+g) / (r-g) using Gordon Growth Model",
        "* Two-phase growth: Years 1-3 near-term, Years 4-5 mature growth",
        "* Weighted Valuation combines Bear/Base/Bull using your probability weights",
        "* Sensitivity tables show impact of key assumption changes"
    ]
    for i, note in enumerate(notes):
        ws[f'A{63+i}'] = note
        ws[f'A{63+i}'].font = Font(size=10, color="666666")

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 30

    # =========================================================================
    # AI INSIGHTS SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis:
        ws_ai = wb.create_sheet("AI Analysis")

        # Title
        ws_ai['A1'] = "AI-POWERED ANALYSIS"
        ws_ai['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_ai['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_ai['A2'].font = Font(italic=True, size=10, color="666666")

        # AI Narrative Summary
        ws_ai['A4'] = "EXECUTIVE SUMMARY"
        ws_ai['A4'].font = section_font
        ws_ai.merge_cells('A5:H7')
        ws_ai['A5'] = ai_analysis.ai_narrative
        ws_ai['A5'].alignment = Alignment(wrap_text=True, vertical='top')

        # Suggested Parameters Section
        ws_ai['A9'] = "AI-SUGGESTED PARAMETERS"
        ws_ai['A9'].font = section_font

        ws_ai['A10'] = "Parameter"
        ws_ai['B10'] = "AI Suggestion"
        ws_ai['C10'] = "Confidence"
        ws_ai['D10'] = "Rationale"
        for col in ['A', 'B', 'C', 'D']:
            ws_ai[f'{col}10'].font = header_font

        ws_ai['A11'] = "Growth Rate (Yr 1-3)"
        ws_ai['B11'] = ai_analysis.suggested_growth_rate
        ws_ai['B11'].number_format = '0.0%'
        ws_ai['B11'].fill = ai_fill
        ws_ai['C11'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D11'] = "Based on historical CAGR with mean reversion"

        ws_ai['A12'] = "Growth Rate (Yr 4-5)"
        ws_ai['B12'] = ai_analysis.suggested_growth_rate * 0.6
        ws_ai['B12'].number_format = '0.0%'
        ws_ai['B12'].fill = ai_fill
        ws_ai['C12'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D12'] = "Mature phase assumes slower growth"

        ws_ai['A13'] = "Terminal Growth Rate"
        ws_ai['B13'] = ai_analysis.suggested_terminal_rate
        ws_ai['B13'].number_format = '0.0%'
        ws_ai['B13'].fill = ai_fill
        ws_ai['C13'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D13'] = "Conservative perpetual decline assumption"

        ws_ai['A14'] = "Discount Rate"
        ws_ai['B14'] = ai_analysis.suggested_discount_rate
        ws_ai['B14'].number_format = '0.0%'
        ws_ai['B14'].fill = ai_fill
        ws_ai['C14'] = f"{ai_analysis.confidence_score:.0%}"
        ws_ai['D14'] = "Risk-adjusted based on volatility"

        # Genre Classification
        ws_ai['A16'] = "GENRE CLASSIFICATION"
        ws_ai['A16'].font = section_font
        ws_ai['A17'] = "Detected/Selected Genre:"
        ws_ai['B17'] = ai_analysis.genre_classification.replace('_', ' ').title()
        ws_ai['B17'].font = Font(bold=True)

        # Benchmark Comparison
        ws_ai['A19'] = "INDUSTRY BENCHMARK COMPARISON"
        ws_ai['A19'].font = section_font

        benchmark = ai_analysis.decay_curve_comparison
        ws_ai['A20'] = "Status:"
        ws_ai['B20'] = benchmark.get('assessment', 'N/A')

        ws_ai['A21'] = "Your Avg Decay Rate:"
        ws_ai['B21'] = benchmark.get('avg_actual_decay', 0)
        ws_ai['B21'].number_format = '0.0%'

        ws_ai['A22'] = "Industry Benchmark:"
        ws_ai['B22'] = benchmark.get('avg_benchmark_decay', 0)
        ws_ai['B22'].number_format = '0.0%'

        ws_ai['A23'] = "Variance:"
        variance = benchmark.get('variance_from_benchmark', 0)
        ws_ai['B23'] = variance
        ws_ai['B23'].number_format = '+0.0%;-0.0%'
        if variance > 0:
            ws_ai['B23'].font = Font(color="006600")  # Green for outperformance
        elif variance < 0:
            ws_ai['B23'].font = Font(color="CC0000")  # Red for underperformance

        # Risk Factors
        ws_ai['A25'] = "RISK FACTORS"
        ws_ai['A25'].font = section_font
        for i, risk in enumerate(ai_analysis.risk_factors[:5]):
            ws_ai[f'A{26+i}'] = f"• {risk}"
            ws_ai[f'A{26+i}'].font = Font(color="CC0000")

        # Opportunities
        ws_ai['A32'] = "OPPORTUNITIES"
        ws_ai['A32'].font = section_font
        for i, opp in enumerate(ai_analysis.opportunities[:4]):
            ws_ai[f'A{33+i}'] = f"• {opp}"
            ws_ai[f'A{33+i}'].font = Font(color="006600")

        # Column widths
        ws_ai.column_dimensions['A'].width = 24
        ws_ai.column_dimensions['B'].width = 16
        ws_ai.column_dimensions['C'].width = 12
        ws_ai.column_dimensions['D'].width = 40
        ws_ai.column_dimensions['E'].width = 14

    # =========================================================================
    # MONTE CARLO SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis and ai_analysis.monte_carlo_results:
        ws_mc = wb.create_sheet("Monte Carlo")
        mc = ai_analysis.monte_carlo_results

        # Title
        ws_mc['A1'] = "MONTE CARLO SIMULATION RESULTS"
        ws_mc['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_mc['A2'] = f"Based on {mc['n_simulations']:,} simulations"
        ws_mc['A2'].font = Font(italic=True, size=10, color="666666")

        # Summary Statistics
        ws_mc['A4'] = "VALUATION DISTRIBUTION"
        ws_mc['A4'].font = section_font

        ws_mc['A5'] = "Statistic"
        ws_mc['B5'] = "Value"
        ws_mc['A5'].font = header_font
        ws_mc['B5'].font = header_font

        ws_mc['A6'] = "Mean Valuation"
        ws_mc['B6'] = mc['mean']
        ws_mc['B6'].number_format = '$#,##0'

        ws_mc['A7'] = "Median Valuation"
        ws_mc['B7'] = mc['median']
        ws_mc['B7'].number_format = '$#,##0'
        ws_mc['B7'].fill = weighted_fill
        ws_mc['B7'].font = Font(bold=True)

        ws_mc['A8'] = "Standard Deviation"
        ws_mc['B8'] = mc['std_dev']
        ws_mc['B8'].number_format = '$#,##0'

        ws_mc['A9'] = "Minimum"
        ws_mc['B9'] = mc['min']
        ws_mc['B9'].number_format = '$#,##0'

        ws_mc['A10'] = "Maximum"
        ws_mc['B10'] = mc['max']
        ws_mc['B10'].number_format = '$#,##0'

        # Percentiles
        ws_mc['A12'] = "PERCENTILE DISTRIBUTION"
        ws_mc['A12'].font = section_font

        ws_mc['A13'] = "Percentile"
        ws_mc['B13'] = "Valuation"
        ws_mc['C13'] = "Interpretation"
        for col in ['A', 'B', 'C']:
            ws_mc[f'{col}13'].font = header_font

        percentiles = mc['percentiles']
        ws_mc['A14'] = "5th (Downside)"
        ws_mc['B14'] = percentiles['p5']
        ws_mc['B14'].number_format = '$#,##0'
        ws_mc['B14'].fill = scenario_bear_fill
        ws_mc['C14'] = "Worst 5% of outcomes"

        ws_mc['A15'] = "10th"
        ws_mc['B15'] = percentiles['p10']
        ws_mc['B15'].number_format = '$#,##0'
        ws_mc['C15'] = "Conservative estimate"

        ws_mc['A16'] = "25th"
        ws_mc['B16'] = percentiles['p25']
        ws_mc['B16'].number_format = '$#,##0'
        ws_mc['C16'] = "Lower quartile"

        ws_mc['A17'] = "50th (Median)"
        ws_mc['B17'] = percentiles['p50']
        ws_mc['B17'].number_format = '$#,##0'
        ws_mc['B17'].fill = scenario_base_fill
        ws_mc['B17'].font = Font(bold=True)
        ws_mc['C17'] = "Most likely outcome"

        ws_mc['A18'] = "75th"
        ws_mc['B18'] = percentiles['p75']
        ws_mc['B18'].number_format = '$#,##0'
        ws_mc['C18'] = "Upper quartile"

        ws_mc['A19'] = "90th"
        ws_mc['B19'] = percentiles['p90']
        ws_mc['B19'].number_format = '$#,##0'
        ws_mc['C19'] = "Optimistic estimate"

        ws_mc['A20'] = "95th (Upside)"
        ws_mc['B20'] = percentiles['p95']
        ws_mc['B20'].number_format = '$#,##0'
        ws_mc['B20'].fill = scenario_bull_fill
        ws_mc['C20'] = "Best 5% of outcomes"

        # Confidence Intervals
        ws_mc['A22'] = "CONFIDENCE INTERVALS"
        ws_mc['A22'].font = section_font

        ws_mc['A23'] = "90% Confidence Range:"
        ws_mc['B23'] = percentiles['p5']
        ws_mc['B23'].number_format = '$#,##0'
        ws_mc['C23'] = "to"
        ws_mc['D23'] = percentiles['p95']
        ws_mc['D23'].number_format = '$#,##0'

        ws_mc['A24'] = "80% Confidence Range:"
        ws_mc['B24'] = percentiles['p10']
        ws_mc['B24'].number_format = '$#,##0'
        ws_mc['C24'] = "to"
        ws_mc['D24'] = percentiles['p90']
        ws_mc['D24'].number_format = '$#,##0'

        ws_mc['A25'] = "50% Confidence Range:"
        ws_mc['B25'] = percentiles['p25']
        ws_mc['B25'].number_format = '$#,##0'
        ws_mc['C25'] = "to"
        ws_mc['D25'] = percentiles['p75']
        ws_mc['D25'].number_format = '$#,##0'

        # Risk Metrics
        ws_mc['A27'] = "RISK METRICS"
        ws_mc['A27'].font = section_font

        ws_mc['A28'] = "Downside Risk (P10)"
        ws_mc['B28'] = mc['downside_risk']
        ws_mc['B28'].number_format = '$#,##0'
        ws_mc['C28'] = "10% chance of being below this"

        ws_mc['A29'] = "Upside Potential (P90)"
        ws_mc['B29'] = mc['upside_potential']
        ws_mc['B29'].number_format = '$#,##0'
        ws_mc['C29'] = "10% chance of exceeding this"

        ws_mc['A30'] = "Value at Risk (vs Median)"
        ws_mc['B30'] = percentiles['p50'] - percentiles['p10']
        ws_mc['B30'].number_format = '$#,##0'
        ws_mc['C30'] = "Potential downside from median"

        # Column widths
        ws_mc.column_dimensions['A'].width = 24
        ws_mc.column_dimensions['B'].width = 16
        ws_mc.column_dimensions['C'].width = 26
        ws_mc.column_dimensions['D'].width = 14

    # =========================================================================
    # DECAY CURVE BENCHMARK SHEET (if AI analysis available)
    # =========================================================================
    if ai_analysis and ai_analysis.decay_curve_comparison:
        ws_decay = wb.create_sheet("Decay Benchmarks")
        benchmark = ai_analysis.decay_curve_comparison

        # Title
        ws_decay['A1'] = "DECAY CURVE BENCHMARK ANALYSIS"
        ws_decay['A1'].font = Font(bold=True, size=16, color="7B68EE")
        ws_decay['A2'] = f"Genre: {benchmark.get('genre', 'Mixed').replace('_', ' ').title()}"
        ws_decay['A2'].font = Font(italic=True, size=11)

        # Genre description
        ws_decay['A3'] = benchmark.get('genre_description', '')
        ws_decay['A3'].font = Font(italic=True, size=10, color="666666")

        # Overall Assessment
        ws_decay['A5'] = "OVERALL ASSESSMENT"
        ws_decay['A5'].font = section_font
        ws_decay['A6'] = benchmark.get('assessment', 'N/A')
        status = benchmark.get('overall_status', 'at_benchmark')
        if status == 'above_benchmark':
            ws_decay['A6'].font = Font(bold=True, color="006600")
        elif status == 'below_benchmark':
            ws_decay['A6'].font = Font(bold=True, color="CC0000")
        else:
            ws_decay['A6'].font = Font(bold=True)

        # Year-by-Year Comparison
        ws_decay['A8'] = "YEAR-BY-YEAR COMPARISON"
        ws_decay['A8'].font = section_font

        ws_decay['A9'] = "Year"
        ws_decay['B9'] = "Your Decay"
        ws_decay['C9'] = "Benchmark"
        ws_decay['D9'] = "Difference"
        ws_decay['E9'] = "Status"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_decay[f'{col}9'].font = header_font

        year_by_year = benchmark.get('year_by_year', [])
        for i, yby in enumerate(year_by_year[:5]):
            row = 10 + i
            ws_decay[f'A{row}'] = f"Year {yby['year']}"
            ws_decay[f'B{row}'] = yby['actual']
            ws_decay[f'B{row}'].number_format = '0.0%'
            ws_decay[f'C{row}'] = yby['benchmark']
            ws_decay[f'C{row}'].number_format = '0.0%'
            ws_decay[f'D{row}'] = yby['difference']
            ws_decay[f'D{row}'].number_format = '+0.0%;-0.0%'
            ws_decay[f'E{row}'] = yby['status'].replace('_', ' ').title()

            if yby['status'] == 'outperforming':
                ws_decay[f'E{row}'].font = Font(color="006600")
            elif yby['status'] == 'underperforming':
                ws_decay[f'E{row}'].font = Font(color="CC0000")

        # Industry Benchmarks Reference
        ws_decay['A17'] = "INDUSTRY DECAY BENCHMARKS BY GENRE"
        ws_decay['A17'].font = section_font

        ws_decay['A18'] = "Genre"
        ws_decay['B18'] = "Year 1"
        ws_decay['C18'] = "Year 2"
        ws_decay['D18'] = "Year 3"
        ws_decay['E18'] = "Year 4"
        ws_decay['F18'] = "Year 5+"
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_decay[f'{col}18'].font = header_font

        row = 19
        for genre, rates in GENRE_DECAY_BENCHMARKS.items():
            ws_decay[f'A{row}'] = genre.replace('_', ' ').title()
            ws_decay[f'B{row}'] = rates['year_1']
            ws_decay[f'C{row}'] = rates['year_2']
            ws_decay[f'D{row}'] = rates['year_3']
            ws_decay[f'E{row}'] = rates['year_4']
            ws_decay[f'F{row}'] = rates['year_5_plus']
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_decay[f'{col}{row}'].number_format = '0%'

            # Highlight selected genre
            if genre == benchmark.get('genre', ''):
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws_decay[f'{col}{row}'].fill = ai_fill
            row += 1

        # Column widths
        ws_decay.column_dimensions['A'].width = 20
        ws_decay.column_dimensions['B'].width = 12
        ws_decay.column_dimensions['C'].width = 12
        ws_decay.column_dimensions['D'].width = 12
        ws_decay.column_dimensions['E'].width = 14
        ws_decay.column_dimensions['F'].width = 12

    # =========================================================================
    # RAW DATA SHEET (if raw data provided)
    # =========================================================================
    if raw_data is not None:
        ws_raw = wb.create_sheet(title="Raw Data")

        # Write headers
        for col_idx, col_name in enumerate(raw_data.columns, 1):
            cell = ws_raw.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        # Write data rows
        for row_idx, row in enumerate(raw_data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_raw.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths based on content
        for col_idx, col_name in enumerate(raw_data.columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, min(len(raw_data) + 2, 100)):  # Sample first 100 rows
                cell_value = ws_raw.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws_raw.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def process_csv(file_storage, enable_ai=True, genre="mixed", n_simulations=1000, oauth_token=None):
    """Process uploaded CSV and return Excel bytes + filename."""

    # Read the file
    filename = file_storage.filename
    if filename.endswith('.xlsx'):
        df = pd.read_excel(file_storage)
    else:
        df = pd.read_csv(file_storage)

    # Find the amount column
    amount_col = None
    for col in ['payable_amount', 'amount', 'earnings', 'royalty']:
        if col in df.columns.str.lower().tolist():
            amount_col = [c for c in df.columns if c.lower() == col][0]
            break

    if amount_col is None:
        amount_cols = [c for c in df.columns if 'amount' in c.lower()]
        if amount_cols:
            amount_col = amount_cols[0]
        else:
            raise ValueError("Could not find an amount/earnings column in the CSV")

    # Find the year column
    year_col = None
    for col in ['distribution_year', 'year', 'date']:
        if col in df.columns.str.lower().tolist():
            year_col = [c for c in df.columns if c.lower() == col][0]
            break

    if year_col is None:
        raise ValueError("Could not find a year column in the CSV")

    # Sum by year
    yearly = df.groupby(year_col)[amount_col].sum().sort_index()

    # Get years
    current_year = datetime.now().year
    years_list = sorted(yearly.index)

    # Extract values
    ytd = yearly.get(current_year, 0)
    year_minus_1 = yearly.get(current_year - 1, 0)
    year_minus_2 = yearly.get(current_year - 2, 0)
    year_minus_3 = yearly.get(current_year - 3, 0)

    # If no current year data, shift
    if ytd == 0 and years_list:
        latest = max(years_list)
        ytd = yearly.get(latest, 0)
        year_minus_1 = yearly.get(latest - 1, 0)
        year_minus_2 = yearly.get(latest - 2, 0)
        year_minus_3 = yearly.get(latest - 3, 0)

    # Base year = most recent full year
    base_year = year_minus_1 if year_minus_1 > 0 else ytd

    # Convert yearly data to dict for AI analysis
    yearly_data = {int(k): float(v) for k, v in yearly.items()}

    # Run AI analysis if enabled
    ai_analysis = None
    if enable_ai and AI_AVAILABLE and base_year > 0:
        try:
            ai_analysis = run_full_analysis(
                yearly_data=yearly_data,
                base_year=base_year,
                genre=genre,
                oauth_token=oauth_token,
                n_simulations=n_simulations
            )
        except Exception as e:
            print(f"AI analysis error: {e}")
            ai_analysis = None

    # Generate output filename
    base_name = os.path.splitext(filename)[0]
    if 'listing' in base_name.lower():
        match = re.search(r'listing[-_]?(\d+)', base_name, re.IGNORECASE)
        if match:
            royalty_name = f"Listing {match.group(1)}"
        else:
            royalty_name = base_name
    else:
        royalty_name = base_name

    output_filename = f"{royalty_name} Valuation.xlsx"

    # Create the valuation
    excel_bytes = create_valuation_template(
        royalty_name=royalty_name,
        year_minus_3=year_minus_3,
        year_minus_2=year_minus_2,
        year_minus_1=year_minus_1,
        ytd=ytd,
        base_year=base_year,
        ai_analysis=ai_analysis,
        yearly_data=yearly_data,
        raw_data=df
    )

    return excel_bytes, output_filename


@app.route('/')
def index():
    token_configured = bool(os.environ.get('ANTHROPIC_OAUTH_TOKEN'))
    return render_template_string(HTML_TEMPLATE, token_configured=token_configured)


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files:
        return 'No file uploaded', 400

    file = request.files['file']
    if file.filename == '':
        return 'No file selected', 400

    try:
        # Get options from form
        enable_ai = request.form.get('enable_ai') == 'on'
        genre = request.form.get('genre', 'mixed')
        n_simulations = int(request.form.get('simulations', 1000))
        oauth_token = request.form.get('oauth_token', '').strip() or os.environ.get('ANTHROPIC_OAUTH_TOKEN') or None

        excel_bytes, output_filename = process_csv(
            file,
            enable_ai=enable_ai,
            genre=genre,
            n_simulations=n_simulations,
            oauth_token=oauth_token
        )

        response = send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
        response.headers['X-Filename'] = output_filename
        return response

    except Exception as e:
        return str(e), 400


if __name__ == '__main__':
    import socket

    # Get local IP for mobile access
    hostname = socket.gethostname()
    try:
        local_ip = socket.gethostbyname(hostname)
    except:
        local_ip = '127.0.0.1'

    print("\n" + "="*50)
    print("  ROYALTY VALUATION TOOL - AI-POWERED VERSION")
    print("="*50)
    print(f"\n  Open in browser:")
    print(f"    - On this computer: http://localhost:5000")
    print(f"    - On your phone:    http://{local_ip}:5000")
    print(f"\n  (Make sure your phone is on the same WiFi)")
    print(f"\n  AI Analysis: {'ENABLED' if AI_AVAILABLE else 'DISABLED (install numpy)'}")
    print("\n  Press Ctrl+C to stop the server")
    print("="*50 + "\n")

    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
